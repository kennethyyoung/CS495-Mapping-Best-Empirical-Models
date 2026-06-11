"""Regenerate phase5_42_cdeotte_timeline.png (Figure 8) — multi-status design.

Canonical logic mirrors notebook 02_reanalysis.ipynb cell "4.2 cdeotte timeline",
but plots cdeotte's roles as INDEPENDENT swimlanes (WIN / COMMENTER / CITED) so a
single competition can appear in more than one lane. The earlier version assigned
each entry one mutually-exclusive status with priority WIN > COMMENTER > CITED,
which collapsed the CITED lane to n=0 — every competition where cdeotte is cited is
also one where he commented, so all 6 citations were absorbed into COMMENTER. The
lane design surfaces those citations honestly (a dot in both COMMENTER and CITED).

Standalone (openpyxl + matplotlib, no pandas) so it can be rerun without the full
notebook environment. After running, sync with: python scripts/sync_report_figures.py
"""

from datetime import datetime
from pathlib import Path

import matplotlib.pyplot as plt
import matplotlib.dates as mdates
import openpyxl

ROOT = Path(__file__).resolve().parents[3]
XLSX = ROOT / "data" / "kaggle_meta_analysis.xlsx"
OUT = ROOT / "analysis" / "figures" / "phase5" / "phase5_42_cdeotte_timeline.png"

HANDLE = "cdeotte"


def load_sheet(wb, name):
    ws = wb[name]
    hdr = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    rows = [dict(zip(hdr, [c.value for c in r])) for r in ws.iter_rows(min_row=2)]
    return [r for r in rows if any(v not in (None, "") for v in r.values())]


def as_date(v):
    if isinstance(v, datetime):
        return v
    return datetime.strptime(str(v).strip(), "%Y-%m-%d")


def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    pa = load_sheet(wb, "Paradigm & Attribution")
    cd = load_sheet(wb, "Competition Data")

    end_date = {
        (str(r["competition_ref"]).strip(), str(r["finish_rank"]).strip()): r["end_date"]
        for r in cd
    }

    # Per-entry independent roles (an entry can hold several).
    wins, commenters, cited, absent = [], [], [], []
    for r in pa:
        key = (str(r["competition_ref"]).strip(), str(r["finish_rank"]).strip())
        d = as_date(end_date[key])
        author = str(r.get("author") or "").lower()
        comm = str(r.get("notable_commenters") or "").lower()
        cite = str(r.get("cited_community_members") or "").lower()
        hit = False
        if HANDLE in author:
            wins.append(d); hit = True
        if HANDLE in comm:
            commenters.append(d); hit = True
        if HANDLE in cite:
            cited.append(d); hit = True
        if not hit:
            absent.append(d)

    # Lanes: WIN (top) -> COMMENTER -> CITED -> absent baseline (bottom).
    Y = {"WIN": 3, "COMMENTER": 2, "CITED": 1, "ABSENT": 0}
    color = {"WIN": "#d62728", "COMMENTER": "#7f7f7f", "CITED": "#1f77b4"}
    size = {"WIN": 180, "COMMENTER": 70, "CITED": 90}

    fig, ax = plt.subplots(figsize=(10, 3.8))

    # Faint absent ticks for timeline density context.
    ax.scatter(absent, [Y["ABSENT"]] * len(absent), s=15, color="lightgray",
               alpha=0.45, zorder=1, label=f"absent (n={len(absent)})")
    for role, pts in (("CITED", cited), ("COMMENTER", commenters), ("WIN", wins)):
        ax.scatter(pts, [Y[role]] * len(pts), s=size[role], color=color[role],
                   alpha=0.85, edgecolor="white", linewidth=1.2, zorder=3,
                   label=f"{role} (n={len(pts)})")

    # First-win and first-presence markers (presence = earliest of any role).
    first_win = min(wins)
    first_presence = min(wins + commenters + cited)
    ax.axvline(first_win, linestyle="--", color="#d62728", alpha=0.4)
    ax.annotate(f'First win: {first_win.strftime("%b %Y")} (s4e12)',
                xy=(first_win, Y["WIN"]), xytext=(first_win, Y["WIN"] + 0.55),
                ha="center", fontsize=9, color="#d62728")
    ax.axvline(first_presence, linestyle=":", color="gray", alpha=0.4)
    ax.annotate(f'First documented presence:\n{first_presence.strftime("%b %Y")} (s3e3 commenter)',
                xy=(first_presence, Y["ABSENT"]), xytext=(first_presence, Y["ABSENT"] - 0.5),
                ha="center", va="top", fontsize=9, color="gray")

    ax.set_yticks([Y["WIN"], Y["COMMENTER"], Y["CITED"], Y["ABSENT"]])
    ax.set_yticklabels(["wins", "commenter", "cited", "absent"])
    ax.set_ylim(Y["ABSENT"] - 1.4, Y["WIN"] + 1.0)
    ax.set_xlabel("Competition end date")
    ax.set_title("cdeotte timeline: observer (2023) → dominant winner (2024–2026)", pad=18)
    ax.xaxis.set_major_locator(mdates.YearLocator())
    ax.xaxis.set_major_formatter(mdates.DateFormatter("%Y"))
    ax.legend(loc="center left", bbox_to_anchor=(1.0, 0.5), frameon=False, fontsize=9)
    for sp in ("top", "right"):
        ax.spines[sp].set_visible(False)
    plt.tight_layout()
    plt.savefig(OUT, bbox_inches="tight", dpi=150)
    print(f"wrote {OUT.relative_to(ROOT)}")
    print(f"  wins={len(wins)} commenter={len(commenters)} cited={len(cited)} absent={len(absent)}")


if __name__ == "__main__":
    main()

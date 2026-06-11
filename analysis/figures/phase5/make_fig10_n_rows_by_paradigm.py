"""Regenerate phase5_24_paradigm_n_rows.png (Figure 10).

Mirrors notebook 02_reanalysis.ipynb cell "3. dataset size per winning entry",
minus the 50K vertical reference line. That line marked cdeotte's "small data"
heuristic (coupling C5), which the corpus contradicts (23% support, Section 4.3)
and which this figure's own discussion (Section 4.8) never invokes -- it only
cluttered the axis (colliding with the 100K tick) for no analytical payoff.

Standalone (openpyxl + matplotlib, no pandas). After running, sync with:
python scripts/sync_report_figures.py
"""

from collections import defaultdict
from pathlib import Path

import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import numpy as np
import openpyxl

ROOT = Path(__file__).resolve().parents[3]
XLSX = ROOT / "data" / "kaggle_meta_analysis.xlsx"
OUT = ROOT / "analysis" / "figures" / "phase5" / "phase5_24_paradigm_n_rows.png"

PARADIGM_ORDER = ["ensemble-stacking", "single-model-FE", "lookup-exploit",
                  "problem-fit-NN", "community-template-tweak", "mixed"]
PARADIGM_COLORS = {
    "ensemble-stacking":        "#1f77b4",
    "single-model-FE":          "#ff7f0e",
    "lookup-exploit":           "#d62728",
    "problem-fit-NN":           "#2ca02c",
    "community-template-tweak": "#9467bd",
    "mixed":                    "#7f7f7f",
}


def load_sheet(wb, name):
    ws = wb[name]
    hdr = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    rows = [dict(zip(hdr, [c.value for c in r])) for r in ws.iter_rows(min_row=2)]
    return [r for r in rows if any(v not in (None, "") for v in r.values())]


def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    pa = load_sheet(wb, "Paradigm & Attribution")
    cd = load_sheet(wb, "Competition Data")
    nrows = {
        (str(r["competition_ref"]).strip(), str(r["finish_rank"]).strip()): r["n_rows"]
        for r in cd
    }
    by_paradigm = defaultdict(list)
    for r in pa:
        key = (str(r["competition_ref"]).strip(), str(r["finish_rank"]).strip())
        n = nrows.get(key)
        if n in (None, ""):
            continue
        by_paradigm[str(r["paradigm"]).strip()].append(n)

    plt.rcParams.update({"axes.spines.top": False, "axes.spines.right": False})
    fig, ax = plt.subplots(figsize=(9, 3.8))
    np.random.seed(0)
    for i, paradigm in enumerate(PARADIGM_ORDER):
        sub = by_paradigm.get(paradigm, [])
        if not sub:
            continue
        y_jitter = np.full(len(sub), i) + (np.random.rand(len(sub)) - 0.5) * 0.3
        ax.scatter(sub, y_jitter, s=70, color=PARADIGM_COLORS[paradigm],
                   edgecolor="white", linewidth=0.8, alpha=0.85, zorder=3)
    ax.set_xscale("log")
    ax.set_yticks(range(len(PARADIGM_ORDER)))
    ax.set_yticklabels(PARADIGM_ORDER)
    ax.invert_yaxis()
    ax.set_xlabel("Training rows (log scale)")
    ax.set_title("Dataset size (n_rows) per winning entry, grouped by paradigm")
    ax.xaxis.set_major_formatter(mticker.FuncFormatter(
        lambda x, _: f"{int(x/1e6)}M" if x >= 1e6 else (f"{int(x/1e3)}K" if x >= 1e3 else str(int(x)))))
    ax.grid(axis="x", alpha=0.3)
    plt.tight_layout()
    plt.savefig(OUT, bbox_inches="tight", dpi=150)
    print(f"wrote {OUT.relative_to(ROOT)}")


if __name__ == "__main__":
    main()

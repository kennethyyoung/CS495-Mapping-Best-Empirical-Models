"""Check that finish_rank matches the highest-ranked source available in writeups/."""
import openpyxl
from pathlib import Path

ROOT = Path(__file__).parent.parent
XL_PATH = ROOT / "data" / "kaggle_meta_analysis.xlsx"
WRITEUPS = ROOT / "data" / "writeups"

wb = openpyxl.load_workbook(XL_PATH)
ws = wb["Competition Data"]
headers = [c.value for c in ws[1]]
ref_col  = headers.index("competition_ref") + 1
rank_col = headers.index("finish_rank") + 1

# Map rank keywords in filenames
RANK_KEYWORDS = {1: ["1st", "first"], 2: ["2nd", "second"], 3: ["3rd", "third"]}

print(f"{'Competition':<40} {'finish_rank':>12}  {'files found'}")
print("-" * 90)

for comp_dir in sorted(WRITEUPS.iterdir()):
    if not comp_dir.is_dir():
        continue
    slug = comp_dir.name

    # Get finish_rank from Excel
    finish_rank = None
    for r in range(2, ws.max_row + 1):
        if str(ws.cell(r, ref_col).value) == slug:
            finish_rank = ws.cell(r, rank_col).value
            break

    files = [f.name for f in comp_dir.iterdir()]
    notebooks = [f for f in files if f.endswith(".ipynb")]
    writeups  = [f for f in files if f.endswith((".txt", ".htm"))]

    # Determine what ranks are present
    ranks_present = {}
    for f in files:
        fname_lower = f.lower()
        for rank, keywords in RANK_KEYWORDS.items():
            if any(k in fname_lower for k in keywords):
                ranks_present.setdefault(rank, []).append(f)

    # Highest rank with a notebook
    nb_ranks = []
    for f in notebooks:
        fname_lower = f.lower()
        for rank, keywords in RANK_KEYWORDS.items():
            if any(k in fname_lower for k in keywords):
                nb_ranks.append(rank)
    highest_nb_rank = min(nb_ranks) if nb_ranks else None

    # Flag mismatches
    flag = ""
    if highest_nb_rank and finish_rank != highest_nb_rank:
        flag = f"  *** MISMATCH: notebook rank={highest_nb_rank}, finish_rank={finish_rank}"
    elif not highest_nb_rank and writeups:
        # No notebooks — check if finish_rank matches highest writeup rank
        wu_ranks = []
        for f in writeups:
            fname_lower = f.lower()
            for rank, keywords in RANK_KEYWORDS.items():
                if any(k in fname_lower for k in keywords):
                    wu_ranks.append(rank)
        highest_wu_rank = min(wu_ranks) if wu_ranks else None
        if highest_wu_rank and finish_rank != highest_wu_rank:
            flag = f"  *** MISMATCH: highest writeup rank={highest_wu_rank}, finish_rank={finish_rank}"

    nb_str = f"{len(notebooks)} notebooks, {len(writeups)} writeups"
    print(f"  {slug:<38} rank={finish_rank}   {nb_str}{flag}")

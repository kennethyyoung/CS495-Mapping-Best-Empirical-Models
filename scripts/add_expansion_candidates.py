"""
Adds TPS 2022 + PS S6 expansion candidates to kaggle_candidates_v2.xlsx.
- Adds a 'batch' column (blank for original rows, 'expansion' for new ones)
- Fetches metadata (title, n_teams, end_date, eval_metric) from Kaggle API
- URL is the competition page, matching existing entries
- Skips any slug already present
"""
import json
import time
from datetime import datetime
from pathlib import Path
import openpyxl
from kaggle import KaggleApi

ROOT = Path(__file__).parent.parent
CAND_PATH = ROOT / "data" / "kaggle_candidates_v2.xlsx"
RAW_DIR   = ROOT / "data" / "raw"

# Screened candidates (aug-2022 excluded — no sources)
NEW_SLUGS = [
    "tabular-playground-series-feb-2022",
    "tabular-playground-series-mar-2022",
    "tabular-playground-series-apr-2022",
    "tabular-playground-series-may-2022",
    "tabular-playground-series-jun-2022",
    "tabular-playground-series-oct-2022",
    "tabular-playground-series-nov-2022",
    "playground-series-s6e1",
    "playground-series-s6e2",
    "playground-series-s6e3",
    "playground-series-s6e4",
]

# Metadata we already have from the discovery run (n_teams, metric, deadline)
# Filled from API output; titles fetched live below
KNOWN_META = {
    "tabular-playground-series-feb-2022": {"n_teams": 1255, "eval_metric": "Categorization Accuracy",  "end_date": "2022-02-28"},
    "tabular-playground-series-mar-2022": {"n_teams": 956,  "eval_metric": "Mean Absolute Error",       "end_date": "2022-03-31"},
    "tabular-playground-series-apr-2022": {"n_teams": 816,  "eval_metric": "Area Under ROC Curve",      "end_date": "2022-04-30"},
    "tabular-playground-series-may-2022": {"n_teams": 1151, "eval_metric": "Area Under ROC Curve",      "end_date": "2022-05-31"},
    "tabular-playground-series-jun-2022": {"n_teams": 844,  "eval_metric": "Root Mean Squared Error",   "end_date": "2022-06-30"},
    "tabular-playground-series-oct-2022": {"n_teams": 463,  "eval_metric": "Mean Columnwise Log Loss",  "end_date": "2022-10-31"},
    "tabular-playground-series-nov-2022": {"n_teams": 689,  "eval_metric": "Log Loss",                  "end_date": "2022-11-30"},
    "playground-series-s6e1":            {"n_teams": 4317, "eval_metric": "Root Mean Squared Error",   "end_date": "2026-01-31"},
    "playground-series-s6e2":            {"n_teams": 4370, "eval_metric": "Roc Auc Score",             "end_date": "2026-02-28"},
    "playground-series-s6e3":            {"n_teams": 4142, "eval_metric": "Roc Auc Score",             "end_date": "2026-03-31"},
    "playground-series-s6e4":            {"n_teams": 4315, "eval_metric": "Balanced Accuracy Score",   "end_date": "2026-04-30"},
}

def get_title(api, slug):
    """Try to fetch competition title from API search."""
    try:
        results = api.competitions_list(search=slug, page=1, page_size=5)
        for c in results.competitions:
            ref = str(c.ref).replace("https://www.kaggle.com/competitions/", "")
            if ref == slug:
                return str(c.title)
    except Exception:
        pass
    # Fallback: format from slug
    return slug.replace("-", " ").title()

def load_n_solution_topics(slug):
    topics_file = RAW_DIR / slug / "topics.json"
    if not topics_file.exists():
        return False, 0
    import re
    SOLUTION_KEYWORDS = re.compile(
        r"\b(1st|2nd|3rd|gold|silver|bronze|winner|winning|solution|writeup|"
        r"place solution|my approach|what worked)\b", re.IGNORECASE
    )
    try:
        topics = json.loads(topics_file.read_text())
        sol = [t for t in topics if SOLUTION_KEYWORDS.search(t.get("title", ""))]
        return len(sol) > 0, len(sol)
    except Exception:
        return False, 0

api = KaggleApi()
api.authenticate()

# Load workbook
wb = openpyxl.load_workbook(CAND_PATH)
ws = wb.active
headers = [c.value for c in ws[1]]

# Add 'batch' column if not present
if "batch" not in headers:
    batch_col = len(headers) + 1
    ws.cell(1, batch_col).value = "batch"
    headers.append("batch")
    print("Added 'batch' column")
else:
    batch_col = headers.index("batch") + 1
    print("'batch' column already exists")

# Collect existing slugs
ref_col = headers.index("competition_ref") + 1
existing = set(
    str(ws.cell(r, ref_col).value)
    for r in range(2, ws.max_row + 1)
)

# Append new rows
added = 0
for slug in NEW_SLUGS:
    if slug in existing:
        print(f"SKIP (already present): {slug}")
        continue

    meta = KNOWN_META[slug]
    title = get_title(api, slug)
    has_sol, n_sol = load_n_solution_topics(slug)
    time.sleep(0.5)

    category = "playground" if slug.startswith("playground-series") else "featured"
    url = f"https://www.kaggle.com/competitions/{slug}"
    end_date = datetime.strptime(meta["end_date"], "%Y-%m-%d")

    row_data = {
        "competition_ref":   slug,
        "title":             title,
        "category":          category,
        "eval_metric":       meta["eval_metric"],
        "n_teams":           meta["n_teams"],
        "end_date":          end_date,
        "is_monetized":      False,
        "has_solution_topic": has_sol,
        "n_solution_topics": n_sol,
        "url":               url,
        "recommended_action": "keep",
        "reason":            "Expansion batch — TPS 2022 / PS S6 tabular competition",
        "batch":             "expansion",
    }

    new_row = [row_data.get(h) for h in headers]
    ws.append(new_row)
    added += 1
    print(f"Added: {slug} | title='{title}' | n_sol_topics={n_sol}")

wb.save(CAND_PATH)
print(f"\nDone. Added {added} new rows. Total rows now: {ws.max_row - 1}")

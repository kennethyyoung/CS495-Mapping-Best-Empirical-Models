"""One-off patch: update specific fields for an existing entry."""
import openpyxl
from pathlib import Path

XL_PATH = Path(__file__).parent.parent / "data" / "kaggle_meta_analysis.xlsx"
SLUG = "tabular-playground-series-feb-2022"
UPDATES = {"dominant_base_model": "other"}  # best_single_model=RadiusNeighborsClassifier

wb = openpyxl.load_workbook(XL_PATH)
ws = wb["Competition Data"]
headers = [cell.value for cell in ws[1]]
col = {h: i + 1 for i, h in enumerate(headers)}

for r in range(2, ws.max_row + 1):
    if ws.cell(r, col["competition_ref"]).value == SLUG:
        for field, value in UPDATES.items():
            ws.cell(r, col[field]).value = value
        wb.save(XL_PATH)
        print(f"Patched row {r} for {SLUG}: {list(UPDATES.keys())}")
        break
else:
    print(f"Slug {SLUG} not found.")

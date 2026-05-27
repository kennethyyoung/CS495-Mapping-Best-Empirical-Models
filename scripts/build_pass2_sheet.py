"""Build Pass 2 'Paradigm & Attribution' sheet in kaggle_meta_analysis.xlsx.

Adds a new sheet capturing what the per-writeup re-evaluation surfaced:
paradigm, origination score, attribution (cited members, public-notebook forks,
academic papers, canonized techniques), author centrality, notable commenters,
and a one-line unique edge per winner.

Also extends the Codebook sheet with definitions for the new fields.
"""

import sys
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

sys.stdout.reconfigure(encoding='utf-8')

XLSX_PATH = Path("data/kaggle_meta_analysis.xlsx")
SHEET_NAME = "Paradigm & Attribution"

HEADERS = [
    "competition_ref", "finish_rank", "author", "paradigm",
    "lookup_exploit_subtype", "lookup_material_present",
    "origination_score", "cited_community_members", "n_cited_members",
    "forked_from_public_notebook", "academic_papers_cited",
    "uses_canonized_technique", "author_surprise_flag",
    "author_recurring_in_set", "author_appearance_count",
    "notable_commenters", "winner_unique_edge", "reeval_doc",
]

CODEBOOK_ADDITIONS = [
    ("paradigm", "ensemble-stacking | single-model-FE | lookup-exploit | problem-fit-NN | community-template-tweak | mixed. Primary winning paradigm per per-writeup re-evaluation."),
    ("lookup_exploit_subtype", "identity | inversion | distance-via-generator-flaw | exact-solution-feature | none. Only meaningful when paradigm = lookup-exploit; otherwise 'none'."),
    ("lookup_material_present", "TRUE | FALSE. TRUE if lookup-able material (exact-match, snap, KDTree, original target lookup, etc.) appeared anywhere in the solution, independent of primary paradigm."),
    ("origination_score", "0-3 ordinal. 0=pure fork; 1=fork+small tweak; 2=significant original work on top of community foundation; 3=mostly original (pipeline+FE+models+ensemble all from winner)."),
    ("cited_community_members", "Semicolon-separated Kaggle handles named in the writeup as technique/notebook/discussion sources."),
    ("n_cited_members", "Count derived from cited_community_members."),
    ("forked_from_public_notebook", "TRUE | FALSE. TRUE if the winning pipeline is built on a single forked public notebook as its base."),
    ("academic_papers_cited", "Semicolon-separated short labels (e.g. 'TFT (1912.09363)', 'PLE (2203.05556)')."),
    ("uses_canonized_technique", "Semicolon-separated from {HC, brute-force-FE, Ridge-as-stacker, RAPIDS-XGB, AG-as-ensembler, target-encoding-stack, LAD-as-stacker, original-as-columns, MLP-stacker, DAE-as-base-encoder, pseudo-labeling, adversarial-validation}. Excluded: Optuna, OOF-stacking-generic, GroupKFold, standard model families."),
    ("author_surprise_flag", "TRUE | FALSE. TRUE if writeup contains explicit 'didn't expect / how on earth / unexpected' framing."),
    ("author_recurring_in_set", "TRUE | FALSE. TRUE if author appears in >1 row of this set (winner or cited source)."),
    ("author_appearance_count", "Total appearances of author across set (wins + citations)."),
    ("notable_commenters", "Semicolon-separated handles for commenters who are EITHER a Kaggle Grandmaster OR appear elsewhere in our set as winner/cited-source/recurring-author."),
    ("winner_unique_edge", "One-line (<=200 chars) description of the single thing distinguishing the winner from the public template / runner-up."),
    ("reeval_doc", "Relative path to per-writeup re-evaluation doc."),
]


ROWS = [
    # icr - room722 - problem-fit-NN
    dict(
        competition_ref="icr-identify-age-related-conditions", finish_rank=1, author="room722",
        paradigm="problem-fit-NN", lookup_exploit_subtype="none", lookup_material_present="FALSE",
        origination_score=2,
        cited_community_members="SAMUEL/muelsamu",
        forked_from_public_notebook="FALSE",
        academic_papers_cited="TFT (1912.09363)",
        uses_canonized_technique="",
        author_surprise_flag="TRUE", author_recurring_in_set="FALSE", author_appearance_count=1,
        notable_commenters="cdeotte;cpmpml;ravi20076",
        winner_unique_edge="NN-only at 617 rows via Variable Selection Network + dropout 0.75 + 10x10-30 repeated CV + multi-label CV + probability reweighting",
        reeval_doc="analysis/writeup-reevaluation/icr_rank1.md",
    ),
    # s4e7 - Cross Sellers (ravi20076+arunklenin) - ensemble-stacking
    dict(
        competition_ref="playground-series-s4e7", finish_rank=1, author="ravi20076+arunklenin (Cross Sellers)",
        paradigm="ensemble-stacking", lookup_exploit_subtype="none", lookup_material_present="TRUE",
        origination_score=3,
        cited_community_members="rzatemizel;uryednap;tilii7;optimistix",
        forked_from_public_notebook="FALSE",
        academic_papers_cited="",
        uses_canonized_technique="brute-force-FE;target-encoding-stack",
        author_surprise_flag="FALSE", author_recurring_in_set="TRUE", author_appearance_count=6,
        notable_commenters="tilii7;optimistix",
        winner_unique_edge="3-stage stacking with XGB final stacker; 78 weak learners; 12-version feature store; component models on data subsets; target reversal post-processing",
        reeval_doc="analysis/writeup-reevaluation/s4e7_rank1.md",
    ),
    # s5e3 rank2 - cdeotte - mixed (small-data, mean blend of 3, kernel SVM)
    dict(
        competition_ref="playground-series-s5e3", finish_rank=2, author="cdeotte",
        paradigm="mixed", lookup_exploit_subtype="none", lookup_material_present="FALSE",
        origination_score=3,
        cited_community_members="",
        forked_from_public_notebook="FALSE",
        academic_papers_cited="",
        uses_canonized_technique="RAPIDS-XGB;original-as-columns",
        author_surprise_flag="FALSE", author_recurring_in_set="TRUE", author_appearance_count=8,
        notable_commenters="",
        winner_unique_edge="Tiny-data playbook: no FE (curse of dimensionality), equal-weight mean blend of 3 models (XGB+TabPFN+RAPIDS SVC poly), Group K Fold by year",
        reeval_doc="analysis/writeup-reevaluation/s5e3_rank2.md",
    ),
    # s5e5 - cdeotte - ensemble-stacking (canonizes GPU HC)
    dict(
        competition_ref="playground-series-s5e5", finish_rank=1, author="cdeotte",
        paradigm="ensemble-stacking", lookup_exploit_subtype="none", lookup_material_present="FALSE",
        origination_score=3,
        cited_community_members="",
        forked_from_public_notebook="FALSE",
        academic_papers_cited="",
        uses_canonized_technique="HC;RAPIDS-XGB;target-encoding-stack",
        author_surprise_flag="FALSE", author_recurring_in_set="TRUE", author_appearance_count=8,
        notable_commenters="mahog",
        winner_unique_edge="GPU Hill Climbing canonization here; 7-model HC ensemble; residual modeling chain (LR->NN->XGB); groupby z-score features (26)",
        reeval_doc="analysis/writeup-reevaluation/s5e5_rank1.md",
    ),
    # s5e12 - wind1234it - ensemble-stacking (distribution-shift, post-cutoff CV)
    dict(
        competition_ref="playground-series-s5e12", finish_rank=1, author="wind1234it",
        paradigm="ensemble-stacking", lookup_exploit_subtype="none", lookup_material_present="TRUE",
        origination_score=3,
        cited_community_members="masayakawamata;daylighth;laureanoarcanio;tilii7;mikhailnaumov;siukeitin",
        forked_from_public_notebook="FALSE",
        academic_papers_cited="",
        uses_canonized_technique="HC;Ridge-as-stacker;original-as-columns",
        author_surprise_flag="FALSE", author_recurring_in_set="FALSE", author_appearance_count=1,
        notable_commenters="masayakawamata;tilii7",
        winner_unique_edge="Post-cutoff CV (cutoff_id=678260) + HC with negative weights + Ridge stacker on rank-transformed top-36; ~100 base OOFs; anti-greedy submission",
        reeval_doc="analysis/writeup-reevaluation/s5e12_rank1.md",
    ),
    # s5e10 - tilii7 - ensemble-stacking (GP features at ensemble stage)
    dict(
        competition_ref="playground-series-s5e10", finish_rank=1, author="tilii7",
        paradigm="ensemble-stacking", lookup_exploit_subtype="none", lookup_material_present="FALSE",
        origination_score=3,
        cited_community_members="masayakawamata;cdeotte;siukeitin;optimistix;mikhailnaumov;mahoganybuttstrings",
        forked_from_public_notebook="FALSE",
        academic_papers_cited="",
        uses_canonized_technique="HC;MLP-stacker;DAE-as-base-encoder",
        author_surprise_flag="FALSE", author_recurring_in_set="TRUE", author_appearance_count=6,
        notable_commenters="cdeotte;masayakawamata",
        winner_unique_edge="11 genetic programming features added at ENSEMBLE stage (not base); Lasso used to reverse-engineer generator formula; Keras NN + CatBoost intermediate stackers",
        reeval_doc="analysis/writeup-reevaluation/s5e10_rank1.md",
    ),
    # s4e10 rank2 - omidbaghchehsaraei - ensemble-stacking (HC, paddykb all-as-cat fork)
    dict(
        competition_ref="playground-series-s4e10", finish_rank=2, author="omidbaghchehsaraei",
        paradigm="ensemble-stacking", lookup_exploit_subtype="none", lookup_material_present="FALSE",
        origination_score=2,
        cited_community_members="paddykb;siukeitin",
        forked_from_public_notebook="TRUE",
        academic_papers_cited="",
        uses_canonized_technique="HC",
        author_surprise_flag="FALSE", author_recurring_in_set="TRUE", author_appearance_count=3,
        notable_commenters="ravaghi;siukeitin",
        winner_unique_edge="HC (hillclimbers library) over 21-24 models including paddykb's all-as-categorical port; anti-greedy submission (lower public LB, higher CV)",
        reeval_doc="analysis/writeup-reevaluation/s4e10_rank2.md",
    ),
    # s5e11 - mahog - ensemble-stacking, diversity-no-tuning
    dict(
        competition_ref="playground-series-s5e11", finish_rank=1, author="mahoganybuttstrings",
        paradigm="ensemble-stacking", lookup_exploit_subtype="none", lookup_material_present="FALSE",
        origination_score=3,
        cited_community_members="masayakawamata;cdeotte;yekenot;yeoyunsianggeremie;jklol86;mikhailnaumov",
        forked_from_public_notebook="FALSE",
        academic_papers_cited="",
        uses_canonized_technique="HC;Ridge-as-stacker;target-encoding-stack",
        author_surprise_flag="FALSE", author_recurring_in_set="TRUE", author_appearance_count=8,
        notable_commenters="cdeotte;tilii7",
        winner_unique_edge="Diversity-without-tuning: 23 model families with default HP + digit interaction combinatorics + TE with alternative targets (employment_status, debt_to_income)",
        reeval_doc="analysis/writeup-reevaluation/s5e11_rank1.md",
    ),
    # s4e1 rank3 - iqbalsyahakbar - ensemble-stacking (TF-IDF+SVD novel)
    dict(
        competition_ref="playground-series-s4e1", finish_rank=3, author="iqbalsyahakbar",
        paradigm="ensemble-stacking", lookup_exploit_subtype="none", lookup_material_present="FALSE",
        origination_score=3,
        cited_community_members="arunklenin;aspillai;paddykb",
        forked_from_public_notebook="FALSE",
        academic_papers_cited="",
        uses_canonized_technique="Ridge-as-stacker;target-encoding-stack",
        author_surprise_flag="FALSE", author_recurring_in_set="FALSE", author_appearance_count=1,
        notable_commenters="paddykb;aspillai",
        winner_unique_edge="TF-IDF+SVD on text-like Surname then encode SVD components via CatBoost (double-layer FE); 30-fold CV for submission; CatBoost-encoder order-sensitivity",
        reeval_doc="analysis/writeup-reevaluation/s4e1_rank3.md",
    ),
    # s5e2 - cdeotte - single-model-FE (single XGB, 500 features)
    dict(
        competition_ref="playground-series-s5e2", finish_rank=1, author="cdeotte",
        paradigm="single-model-FE", lookup_exploit_subtype="none", lookup_material_present="FALSE",
        origination_score=3,
        cited_community_members="jordanbarker",
        forked_from_public_notebook="FALSE",
        academic_papers_cited="",
        uses_canonized_technique="brute-force-FE;target-encoding-stack;RAPIDS-XGB;original-as-columns",
        author_surprise_flag="FALSE", author_recurring_in_set="TRUE", author_appearance_count=8,
        notable_commenters="",
        winner_unique_edge="Single XGB with 500 engineered features on 4M rows; histogram-binned groupby aggs (cdeotte invention); NaN base-2 encoding; nested-fold target-aware FE",
        reeval_doc="analysis/writeup-reevaluation/s5e2_rank1.md",
    ),
    # s5e8 rank2 - mahog - ensemble-stacking with CatBoost-as-stacker (counter-evidence)
    dict(
        competition_ref="playground-series-s5e8", finish_rank=2, author="mahoganybuttstrings",
        paradigm="ensemble-stacking", lookup_exploit_subtype="none", lookup_material_present="FALSE",
        origination_score=3,
        cited_community_members="cdeotte;omidbaghchehsaraei;yekenot;itasps;siukeitin;tilii7",
        forked_from_public_notebook="FALSE",
        academic_papers_cited="",
        uses_canonized_technique="HC;target-encoding-stack",
        author_surprise_flag="FALSE", author_recurring_in_set="TRUE", author_appearance_count=8,
        notable_commenters="cdeotte;tilii7",
        winner_unique_edge="CatBoost as final stacker (beat Ridge by 0.00010 on private) in head-to-head 4-stacker test; 13 model architectures; TE on bigrams using comp + original targets",
        reeval_doc="analysis/writeup-reevaluation/s5e8_rank2.md",
    ),
    # s5e4 rank2 - greysky - single-model-FE (minimalist single LGBM 1552 features)
    dict(
        competition_ref="playground-series-s5e4", finish_rank=2, author="greysky",
        paradigm="single-model-FE", lookup_exploit_subtype="none", lookup_material_present="FALSE",
        origination_score=3,
        cited_community_members="",
        forked_from_public_notebook="FALSE",
        academic_papers_cited="",
        uses_canonized_technique="target-encoding-stack",
        author_surprise_flag="FALSE", author_recurring_in_set="TRUE", author_appearance_count=2,
        notable_commenters="cdeotte",
        winner_unique_edge="Single LightGBM minimalist: 1552 features via 6-way TE combinatorics + descriptive stats; no CV, no ensemble, no HP tuning; Kaggle parallel-notebook experimentation",
        reeval_doc="analysis/writeup-reevaluation/s5e4_rank2.md",
    ),
    # s4e9 - Mart Preusse - ensemble-stacking with NN-stacker, surprised
    dict(
        competition_ref="playground-series-s4e9", finish_rank=1, author="martinapreusse",
        paradigm="ensemble-stacking", lookup_exploit_subtype="none", lookup_material_present="FALSE",
        origination_score=2,
        cited_community_members="yekenot;siukeitin;tilii7;roberthatch;noodl35;cdeotte",
        forked_from_public_notebook="TRUE",
        academic_papers_cited="",
        uses_canonized_technique="MLP-stacker;Ridge-as-stacker",
        author_surprise_flag="TRUE", author_recurring_in_set="FALSE", author_appearance_count=1,
        notable_commenters="cdeotte",
        winner_unique_edge="Last-day fork of yekenot's DeepTables NN starter + added 4 OOFs as features = NN-as-stacker won; outlier classification via IQR; median TE",
        reeval_doc="analysis/writeup-reevaluation/s4e9_rank1.md",
    ),
    # s4e5 - adaubas - ensemble-stacking, EDA-driven (Poisson sum)
    dict(
        competition_ref="playground-series-s4e5", finish_rank=1, author="adaubas (aldparis)",
        paradigm="ensemble-stacking", lookup_exploit_subtype="none", lookup_material_present="FALSE",
        origination_score=3,
        cited_community_members="mfmfmf3;igorvolianiuk;siukeitin;act18l;ambrosm;tilii7",
        forked_from_public_notebook="FALSE",
        academic_papers_cited="",
        uses_canonized_technique="Ridge-as-stacker",
        author_surprise_flag="FALSE", author_recurring_in_set="TRUE", author_appearance_count=4,
        notable_commenters="innixma;tilii7;ambrosm",
        winner_unique_edge="EDA-driven discovery: dataset is Poisson sum -> row-wise stats FE; Ridge with negative weights (positive=False); only 2 submissions in 13 days (trust-CV)",
        reeval_doc="analysis/writeup-reevaluation/s4e5_rank1.md",
    ),
    # s4e11 - ravaghi - ensemble-stacking, AG-as-ensembler
    dict(
        competition_ref="playground-series-s4e11", finish_rank=1, author="ravaghi",
        paradigm="ensemble-stacking", lookup_exploit_subtype="none", lookup_material_present="FALSE",
        origination_score=3,
        cited_community_members="optimistix;tilii7",
        forked_from_public_notebook="FALSE",
        academic_papers_cited="",
        uses_canonized_technique="AG-as-ensembler",
        author_surprise_flag="FALSE", author_recurring_in_set="TRUE", author_appearance_count=3,
        notable_commenters="optimistix;tilii7",
        winner_unique_edge="AutoGluon as final stacker (chosen over HC/Ridge/LR head-to-head); 24 models selected from 69; 4 data pipelines for diversity; anti-greedy submission",
        reeval_doc="analysis/writeup-reevaluation/s4e11_rank1.md",
    ),
    # s5e6 - cdeotte - ensemble-stacking (proto-KGMON, 2268-feature monster)
    dict(
        competition_ref="playground-series-s5e6", finish_rank=1, author="cdeotte",
        paradigm="ensemble-stacking", lookup_exploit_subtype="none", lookup_material_present="TRUE",
        origination_score=2,
        cited_community_members="bizen250;ayushchandramaurya;elainedazzio;ricopue",
        forked_from_public_notebook="FALSE",
        academic_papers_cited="",
        uses_canonized_technique="HC;target-encoding-stack;RAPIDS-XGB;original-as-columns;pseudo-labeling",
        author_surprise_flag="FALSE", author_recurring_in_set="TRUE", author_appearance_count=8,
        notable_commenters="mahog",
        winner_unique_edge="Original-as-columns + cuDF TE at scale: 162 combos x 7 binary targets x 2 datasets = 2268 features; boosting over LR residuals; multi-seed averaging tuned to MAP@K",
        reeval_doc="analysis/writeup-reevaluation/s5e6_rank1.md",
    ),
    # s4e8 - Optimistix - ensemble-stacking (72 OOFs into AG)
    dict(
        competition_ref="playground-series-s4e8", finish_rank=1, author="optimistix",
        paradigm="ensemble-stacking", lookup_exploit_subtype="exact-solution-feature", lookup_material_present="TRUE",
        origination_score=2,
        cited_community_members="ambrosm;siukeitin;nischaydnk;gauravduttakiit;rzatemizel;ravaghi;oscarm524;ravi20076;tilii7;roberthatch;omidbaghchehsaraei;trupologhelper;arunklenin;carlmcbrideellis",
        forked_from_public_notebook="FALSE",
        academic_papers_cited="",
        uses_canonized_technique="HC;AG-as-ensembler;Ridge-as-stacker",
        author_surprise_flag="FALSE", author_recurring_in_set="TRUE", author_appearance_count=6,
        notable_commenters="ravaghi;tilii7;ravi20076",
        winner_unique_edge="Confident-disagreement-overwriting (>0.99 threshold inserts) + 72 OOFs into AutoGluon-as-ensembler; three independent stacker paths converged at 0.98535",
        reeval_doc="analysis/writeup-reevaluation/s4e8_rank1.md",
    ),
    # s4e12 - cdeotte - single-model-FE
    dict(
        competition_ref="playground-series-s4e12", finish_rank=1, author="cdeotte",
        paradigm="single-model-FE", lookup_exploit_subtype="none", lookup_material_present="FALSE",
        origination_score=3,
        cited_community_members="",
        forked_from_public_notebook="FALSE",
        academic_papers_cited="",
        uses_canonized_technique="brute-force-FE;target-encoding-stack;RAPIDS-XGB",
        author_surprise_flag="FALSE", author_recurring_in_set="TRUE", author_appearance_count=8,
        notable_commenters="",
        winner_unique_edge="GPU brute-force search of 145K+ categorical combinations -> kept 170 that moved CV; explicit FE-effectiveness conditioning as strategy axis",
        reeval_doc="analysis/writeup-reevaluation/s4e12_rank1.md",
    ),
    # s4e3 rank2 - Moonlit - community-template-tweak (fork-heavy first-timer)
    dict(
        competition_ref="playground-series-s4e3", finish_rank=2, author="yunqicao (Moonlit)",
        paradigm="community-template-tweak", lookup_exploit_subtype="none", lookup_material_present="FALSE",
        origination_score=1,
        cited_community_members="noepinefrin;arunklenin;cyrilbourgeois;thomasmeiner;ankurgarg04;lucamassaron;arnogils",
        forked_from_public_notebook="TRUE",
        academic_papers_cited="",
        uses_canonized_technique="",
        author_surprise_flag="FALSE", author_recurring_in_set="FALSE", author_appearance_count=1,
        notable_commenters="arunklenin",
        winner_unique_edge="Fork-heavy first-time competitor: 3 OOF ensembles (1 own + 2 public-notebook replications) + Nelder-Mead weights; public-notebook blend alone scores top-10",
        reeval_doc="analysis/writeup-reevaluation/s4e3_rank2.md",
    ),
    # s3e24 rank3 - Ravi - ensemble-stacking, brute-force FE (pre-cdeotte)
    dict(
        competition_ref="playground-series-s3e24", finish_rank=3, author="ravi20076",
        paradigm="ensemble-stacking", lookup_exploit_subtype="none", lookup_material_present="FALSE",
        origination_score=2,
        cited_community_members="arunklenin;cv13j0;oscarm524;paddykb",
        forked_from_public_notebook="FALSE",
        academic_papers_cited="",
        uses_canonized_technique="brute-force-FE",
        author_surprise_flag="FALSE", author_recurring_in_set="TRUE", author_appearance_count=6,
        notable_commenters="paddykb;oscarm524",
        winner_unique_edge="Brute-force FE search at Dec 2023 (predates cdeotte by 12 months, sourced from arunklenin); 8 model families incl TabNet+GAM; 10x1 vs 10x3 CV comparison",
        reeval_doc="analysis/writeup-reevaluation/s3e24_rank3.md",
    ),
    # s3e23 rank2 - oscarm524 - ensemble-stacking, HC (pre-cdeotte)
    dict(
        competition_ref="playground-series-s3e23", finish_rank=2, author="oscarm524",
        paradigm="ensemble-stacking", lookup_exploit_subtype="none", lookup_material_present="FALSE",
        origination_score=2,
        cited_community_members="ambrosm;sauravpandey11",
        forked_from_public_notebook="FALSE",
        academic_papers_cited="",
        uses_canonized_technique="HC",
        author_surprise_flag="FALSE", author_recurring_in_set="TRUE", author_appearance_count=4,
        notable_commenters="ambrosm",
        winner_unique_edge="HC ensemble of 8 models at Nov 2023 (predates cdeotte canonization by 18 months); log-transform on all features; Nystrom LogisticRegression",
        reeval_doc="analysis/writeup-reevaluation/s3e23_rank2.md",
    ),
    # s3e14 - Sergey - lookup-exploit (identity)
    dict(
        competition_ref="playground-series-s3e14", finish_rank=1, author="sergiosaharovskiy",
        paradigm="lookup-exploit", lookup_exploit_subtype="identity", lookup_material_present="TRUE",
        origination_score=1,
        cited_community_members="zhukovoleksiy;paddykb;yzokulu;tetsutani;adaubas;mattop",
        forked_from_public_notebook="TRUE",
        academic_papers_cited="",
        uses_canonized_technique="LAD-as-stacker",
        author_surprise_flag="FALSE", author_recurring_in_set="TRUE", author_appearance_count=3,
        notable_commenters="adaubas;paddykb",
        winner_unique_edge="Match test rows to original dataset by (fruitset, fruitmass) keys; assign exact original yield values (776 unique yields, 2073 test rows); largest 1-2 gap (0.657 MAE)",
        reeval_doc="analysis/writeup-reevaluation/s3e14_rank1.md",
    ),
    # s3e26 rank2 - Hardy Xu - ensemble-stacking with NN-stacker, PLE
    dict(
        competition_ref="playground-series-s3e26", finish_rank=2, author="hardyxu52",
        paradigm="ensemble-stacking", lookup_exploit_subtype="none", lookup_material_present="FALSE",
        origination_score=3,
        cited_community_members="",
        forked_from_public_notebook="FALSE",
        academic_papers_cited="PLE (2203.05556)",
        uses_canonized_technique="MLP-stacker",
        author_surprise_flag="FALSE", author_recurring_in_set="TRUE", author_appearance_count=3,
        notable_commenters="",
        winner_unique_edge="PLE for continuous features in NN; NN-as-final-stacker with per-class probability weights + cross-prediction weight calculation; 10 Optuna sets averaged per GBDT",
        reeval_doc="analysis/writeup-reevaluation/s3e26_rank2.md",
    ),
    # s3e16 rank3 - Ravi - ensemble-stacking, generator augmentation
    dict(
        competition_ref="playground-series-s3e16", finish_rank=3, author="ravi20076",
        paradigm="ensemble-stacking", lookup_exploit_subtype="none", lookup_material_present="TRUE",
        origination_score=3,
        cited_community_members="pandeyg0811;inversion",
        forked_from_public_notebook="FALSE",
        academic_papers_cited="",
        uses_canonized_technique="LAD-as-stacker",
        author_surprise_flag="FALSE", author_recurring_in_set="TRUE", author_appearance_count=6,
        notable_commenters="",
        winner_unique_edge="Generator-driven data augmentation (run official generator notebook with varied params); domain features (meat yield, pseudo BMI); feature-subset diversity per model",
        reeval_doc="analysis/writeup-reevaluation/s3e16_rank3.md",
    ),
    # s3e17 rank3 - ISoft - ensemble-stacking (AutoML stack-of-stacks)
    dict(
        competition_ref="playground-series-s3e17", finish_rank=3, author="mathurinache (ISoft)",
        paradigm="ensemble-stacking", lookup_exploit_subtype="none", lookup_material_present="FALSE",
        origination_score=2,
        cited_community_members="",
        forked_from_public_notebook="FALSE",
        academic_papers_cited="",
        uses_canonized_technique="",
        author_surprise_flag="FALSE", author_recurring_in_set="FALSE", author_appearance_count=1,
        notable_commenters="",
        winner_unique_edge="Pure-AutoML stack-of-stacks: 5 AutoML frameworks (Light AutoML, H2O, FLAML, Lazy Classifier, proprietary Statmining) ensembled; 6h on 4 cores 8GB; 10% public ports",
        reeval_doc="analysis/writeup-reevaluation/s3e17_rank3.md",
    ),
    # s3e11 - ambrosm - ensemble-stacking (Zoo of Models, Ridge stacker canon)
    dict(
        competition_ref="playground-series-s3e11", finish_rank=1, author="ambrosm",
        paradigm="ensemble-stacking", lookup_exploit_subtype="none", lookup_material_present="FALSE",
        origination_score=3,
        cited_community_members="",
        forked_from_public_notebook="FALSE",
        academic_papers_cited="",
        uses_canonized_technique="Ridge-as-stacker;target-encoding-stack",
        author_surprise_flag="FALSE", author_recurring_in_set="TRUE", author_appearance_count=8,
        notable_commenters="sergiosaharovskiy",
        winner_unique_edge="Earliest Ridge-as-stacker (Jun 2023); 18-model Zoo + Ridge weights; duplicate-grouping (360K->3K); PDP-based numerical-as-categorical detection (store_sqft)",
        reeval_doc="analysis/writeup-reevaluation/s3e11_rank1.md",
    ),
    # s3e13 - Umar - ensemble-stacking (mean blend with autoencoder), surprised
    dict(
        competition_ref="playground-series-s3e13", finish_rank=1, author="mufathurrohman (Umar)",
        paradigm="ensemble-stacking", lookup_exploit_subtype="none", lookup_material_present="FALSE",
        origination_score=3,
        cited_community_members="belati;mpwolke",
        forked_from_public_notebook="FALSE",
        academic_papers_cited="",
        uses_canonized_technique="DAE-as-base-encoder",
        author_surprise_flag="TRUE", author_recurring_in_set="FALSE", author_appearance_count=1,
        notable_commenters="",
        winner_unique_edge="First-time competitor: 3-model mean-blend (LGBM + NN + frozen-encoder autoencoder-classifier) at 707 rows; default HP; cites 7-yr-old Porto Seguro autoencoder thread",
        reeval_doc="analysis/writeup-reevaluation/s3e13_rank1.md",
    ),
    # s3e5 - Heitor - single-model-FE (single RAPIDS XGB, trust-CV-no-original)
    dict(
        competition_ref="playground-series-s3e5", finish_rank=1, author="rapela (Heitor)",
        paradigm="single-model-FE", lookup_exploit_subtype="none", lookup_material_present="TRUE",
        origination_score=3,
        cited_community_members="paddykb;abhishek;carlmcbrideellis",
        forked_from_public_notebook="FALSE",
        academic_papers_cited="",
        uses_canonized_technique="RAPIDS-XGB",
        author_surprise_flag="FALSE", author_recurring_in_set="FALSE", author_appearance_count=1,
        notable_commenters="paddykb",
        winner_unique_edge="Earliest single-model winner (Mar 2023): single RAPIDS XGBoost + OptimizedRounder (paddykb) + Optuna + 10-fold CV; chose NOT to use original (CV-driven)",
        reeval_doc="analysis/writeup-reevaluation/s3e5_rank1.md",
    ),
    # s3e10 - seascape - single-model-FE (GAM as primary, R)
    dict(
        competition_ref="playground-series-s3e10", finish_rank=1, author="seascape",
        paradigm="single-model-FE", lookup_exploit_subtype="none", lookup_material_present="FALSE",
        origination_score=3,
        cited_community_members="",
        forked_from_public_notebook="FALSE",
        academic_papers_cited="",
        uses_canonized_technique="",
        author_surprise_flag="FALSE", author_recurring_in_set="FALSE", author_appearance_count=1,
        notable_commenters="paddykb;adaubas",
        winner_unique_edge="GAM as PRIMARY model (R .Rmd); XGB+Lasso predictions DEMOTED to input features for GAM; anti-boosting philosophy; reversed stacker hierarchy",
        reeval_doc="analysis/writeup-reevaluation/s3e10_rank1.md",
    ),
    # s3e9 - ambrosm - ensemble-stacking (zoo, mean blend, anti-Optuna)
    dict(
        competition_ref="playground-series-s3e9", finish_rank=1, author="ambrosm",
        paradigm="ensemble-stacking", lookup_exploit_subtype="none", lookup_material_present="FALSE",
        origination_score=3,
        cited_community_members="",
        forked_from_public_notebook="FALSE",
        academic_papers_cited="",
        uses_canonized_technique="",
        author_surprise_flag="FALSE", author_recurring_in_set="TRUE", author_appearance_count=8,
        notable_commenters="",
        winner_unique_edge="First explicit anti-Optuna stance (manual HP > Optuna); 7-model zoo + mean blend; quantitative trust-CV argument (5407 train vs 721 public LB = coin flip); PDP-based AgeInDays-as-categorical",
        reeval_doc="analysis/writeup-reevaluation/s3e9_rank1.md",
    ),
    # s3e8 rank2 - Craig Thomas - ensemble-stacking (1816 models custom framework)
    dict(
        competition_ref="playground-series-s3e8", finish_rank=2, author="craigmthomas",
        paradigm="ensemble-stacking", lookup_exploit_subtype="none", lookup_material_present="TRUE",
        origination_score=3,
        cited_community_members="",
        forked_from_public_notebook="FALSE",
        academic_papers_cited="",
        uses_canonized_technique="Ridge-as-stacker",
        author_surprise_flag="FALSE", author_recurring_in_set="FALSE", author_appearance_count=1,
        notable_commenters="",
        winner_unique_edge="Custom AutoML-like framework: 1,816 total models (1709 L1 + 105 L2 + Ridge stacker); threshold-based first-level selection; lookup-exploit TESTED and rejected (overfit public LB)",
        reeval_doc="analysis/writeup-reevaluation/s3e8_rank2.md",
    ),
    # s3e6 rank3 - viktortaran - ensemble-stacking (4-GBDT, LGBM stacker)
    dict(
        competition_ref="playground-series-s3e6", finish_rank=3, author="viktortaran",
        paradigm="ensemble-stacking", lookup_exploit_subtype="none", lookup_material_present="FALSE",
        origination_score=3,
        cited_community_members="",
        forked_from_public_notebook="FALSE",
        academic_papers_cited="",
        uses_canonized_technique="",
        author_surprise_flag="FALSE", author_recurring_in_set="FALSE", author_appearance_count=1,
        notable_commenters="",
        winner_unique_edge="4-GBDT stack with LGBM as metamodel (tree-based stacker, counter to Ridge dominance); permutation importance per model; repeated KFold (5x10)",
        reeval_doc="analysis/writeup-reevaluation/s3e6_rank3.md",
    ),
    # s3e1 - Kirderf - ensemble-stacking (AutoGluon + fork FE), surprised
    dict(
        competition_ref="playground-series-s3e1", finish_rank=1, author="kirderf",
        paradigm="ensemble-stacking", lookup_exploit_subtype="none", lookup_material_present="FALSE",
        origination_score=1,
        cited_community_members="dmitryuarov",
        forked_from_public_notebook="TRUE",
        academic_papers_cited="AutoGluon (2003.06505)",
        uses_canonized_technique="AG-as-ensembler",
        author_surprise_flag="TRUE", author_recurring_in_set="FALSE", author_appearance_count=1,
        notable_commenters="innixma",
        winner_unique_edge="Earliest AutoGluon-handles-everything win (Jan 2023); ported dmitryuarov's coordinate-based geo FE + AutoGluon = win; author explicit 'pure AutoML wouldn't have won'",
        reeval_doc="analysis/writeup-reevaluation/s3e1_rank1.md",
    ),
    # s3e7 - Hardy Xu - lookup-exploit (inversion)
    dict(
        competition_ref="playground-series-s3e7", finish_rank=1, author="hardyxu52",
        paradigm="lookup-exploit", lookup_exploit_subtype="inversion", lookup_material_present="TRUE",
        origination_score=3,
        cited_community_members="siukeitin",
        forked_from_public_notebook="FALSE",
        academic_papers_cited="",
        uses_canonized_technique="adversarial-validation",
        author_surprise_flag="FALSE", author_recurring_in_set="TRUE", author_appearance_count=3,
        notable_commenters="siukeitin",
        winner_unique_edge="Opposite-label assignment exploit: 716 test rows had train twins with opposite labels; assigned opposite = +0.014 LB; siukeitin's 'predict 0.5 for test duplicates'",
        reeval_doc="analysis/writeup-reevaluation/s3e7_rank1.md",
    ),
    # s3e3 - Bill Cruise - community-template-tweak (fork + domain FE), surprised
    dict(
        competition_ref="playground-series-s3e3", finish_rank=1, author="bcruise (Bill Cruise)",
        paradigm="community-template-tweak", lookup_exploit_subtype="none", lookup_material_present="FALSE",
        origination_score=1,
        cited_community_members="khawajaabaidullah",
        forked_from_public_notebook="TRUE",
        academic_papers_cited="",
        uses_canonized_technique="",
        author_surprise_flag="TRUE", author_recurring_in_set="FALSE", author_appearance_count=1,
        notable_commenters="cdeotte",
        winner_unique_edge="Forked khawajaabaidullah's 3-GBDT notebook and added 7 domain-driven HR risk-flag features (Age<34, JobHopper, AttritionRisk composite); win was the FE addition",
        reeval_doc="analysis/writeup-reevaluation/s3e3_rank1.md",
    ),
    # s5e7 rank2 - Irfan - mixed (lookup-exploit via match_p + GBDT ensemble)
    dict(
        competition_ref="playground-series-s5e7", finish_rank=2, author="irfanahmad1",
        paradigm="lookup-exploit", lookup_exploit_subtype="exact-solution-feature", lookup_material_present="TRUE",
        origination_score=2,
        cited_community_members="",
        forked_from_public_notebook="FALSE",
        academic_papers_cited="",
        uses_canonized_technique="",
        author_surprise_flag="FALSE", author_recurring_in_set="FALSE", author_appearance_count=1,
        notable_commenters="",
        winner_unique_edge="match_p trick: exact-row-join with public original on 7-feature tuple = match_p feature; 5 GBDT variants + Optuna negative-weight blend with threshold tuning",
        reeval_doc="analysis/writeup-reevaluation/s5e7_rank2.md",
    ),
    # s4e4 - stopwhispering - ensemble-stacking (OpenFE academic tool)
    dict(
        competition_ref="playground-series-s4e4", finish_rank=1, author="stopwhispering",
        paradigm="ensemble-stacking", lookup_exploit_subtype="none", lookup_material_present="FALSE",
        origination_score=3,
        cited_community_members="endofnight17j03;inagana;siukeitin",
        forked_from_public_notebook="FALSE",
        academic_papers_cited="OpenFE (Zhang 2022)",
        uses_canonized_technique="adversarial-validation",
        author_surprise_flag="FALSE", author_recurring_in_set="TRUE", author_appearance_count=3,
        notable_commenters="",
        winner_unique_edge="OpenFE academic tool for automated FE + SFS per model family + 49-model Nelder-Mead with negative coefficients (sum=0.997); anti-NN-for-tabular stance; MLOps",
        reeval_doc="analysis/writeup-reevaluation/s4e4_rank1.md",
    ),
    # s3e4 rank3 - olliekemp - single-model-FE (single CatBoost random-jump ensemble)
    dict(
        competition_ref="playground-series-s3e4", finish_rank=3, author="olliekemp",
        paradigm="single-model-FE", lookup_exploit_subtype="none", lookup_material_present="FALSE",
        origination_score=3,
        cited_community_members="",
        forked_from_public_notebook="FALSE",
        academic_papers_cited="",
        uses_canonized_technique="",
        author_surprise_flag="FALSE", author_recurring_in_set="FALSE", author_appearance_count=1,
        notable_commenters="",
        winner_unique_edge="Custom CatTune class jointly tunes model params + sampling rate + class weights; 500-CatBoost random-jump pseudo-ensemble via Gaussian HP perturbation; aggregate stats over V cols",
        reeval_doc="analysis/writeup-reevaluation/s3e4_rank3.md",
    ),
    # tps-feb-2022 - ambrosm - lookup-exploit (distance-via-generator-flaw)
    dict(
        competition_ref="tabular-playground-series-feb-2022", finish_rank=1, author="ambrosm",
        paradigm="lookup-exploit", lookup_exploit_subtype="distance-via-generator-flaw", lookup_material_present="TRUE",
        origination_score=3,
        cited_community_members="siukeitin",
        forked_from_public_notebook="FALSE",
        academic_papers_cited="Knuth AOCP Vol2",
        uses_canonized_technique="",
        author_surprise_flag="FALSE", author_recurring_in_set="TRUE", author_appearance_count=8,
        notable_commenters="siukeitin",
        winner_unique_edge="Reverse-engineered np.random.RandomState.choice flaw at source-code level (Knuth Alg K); reverse-transform 286 decamer counts -> 100-element sequence; RadiusNeighborsClassifier wins (single model)",
        reeval_doc="analysis/writeup-reevaluation/tps_feb_2022_rank1.md",
    ),
    # tps-may-2022 - The Team - problem-fit-NN (two-branch from xgbfir)
    dict(
        competition_ref="tabular-playground-series-may-2022", finish_rank=1, author="ambrosm+pourchot (The Team)",
        paradigm="problem-fit-NN", lookup_exploit_subtype="none", lookup_material_present="FALSE",
        origination_score=3,
        cited_community_members="sudalairajkumar (SRK);cabaxiom;wti200;pourchot",
        forked_from_public_notebook="TRUE",
        academic_papers_cited="",
        uses_canonized_technique="",
        author_surprise_flag="FALSE", author_recurring_in_set="TRUE", author_appearance_count=8,
        notable_commenters="sudalairajkumar (SRK)",
        winner_unique_edge="Architecture-from-data-structure: xgbfir interaction graph has 2 disconnected components -> built NN with 2 branches mirroring graph; LightGBM interaction_constraints tested",
        reeval_doc="analysis/writeup-reevaluation/tps_may_2022_rank1.md",
    ),
    # tps-jun-2022 - Sebastian - problem-fit-NN (DAE for imputation task)
    dict(
        competition_ref="tabular-playground-series-jun-2022", finish_rank=1, author="sebastianvangerwen",
        paradigm="problem-fit-NN", lookup_exploit_subtype="none", lookup_material_present="FALSE",
        origination_score=3,
        cited_community_members="masatomurakawamm;ehekatlact",
        forked_from_public_notebook="FALSE",
        academic_papers_cited="2106.16057;2002.08338",
        uses_canonized_technique="DAE-as-base-encoder",
        author_surprise_flag="FALSE", author_recurring_in_set="FALSE", author_appearance_count=1,
        notable_commenters="",
        winner_unique_edge="DAE-for-imputation: task IS missing-value imputation; 7-layer MLP with feature-wise mask embeddings (16-dim); masked MSE loss; conditional output (use pred only at masked positions)",
        reeval_doc="analysis/writeup-reevaluation/tps_jun_2022_rank1.md",
    ),
    # s6e1 - mahog - ensemble-stacking (NN-dominant, strength-over-diversity)
    dict(
        competition_ref="playground-series-s6e1", finish_rank=1, author="mahoganybuttstrings",
        paradigm="ensemble-stacking", lookup_exploit_subtype="none", lookup_material_present="FALSE",
        origination_score=3,
        cited_community_members="cdeotte;yekenot;omidbaghchehsaraei;siukeitin;masayakawamata;mikhailnaumov",
        forked_from_public_notebook="FALSE",
        academic_papers_cited="",
        uses_canonized_technique="Ridge-as-stacker;target-encoding-stack;MLP-stacker",
        author_surprise_flag="FALSE", author_recurring_in_set="TRUE", author_appearance_count=8,
        notable_commenters="cdeotte;masayakawamata",
        winner_unique_edge="Strength-over-diversity inversion (vs mahog's prior diversity-no-tuning); WandB sweeps save ALL trial OOFs (190 total); Ridge + Centered Isotonic at TWO points (per-OOF + post-Ridge)",
        reeval_doc="analysis/writeup-reevaluation/s6e1_rank1.md",
    ),
    # s6e2 - masayakawamata - ensemble-stacking with validation discipline headline
    dict(
        competition_ref="playground-series-s6e2", finish_rank=1, author="masayakawamata",
        paradigm="ensemble-stacking", lookup_exploit_subtype="none", lookup_material_present="TRUE",
        origination_score=3,
        cited_community_members="cdeotte;tilii7;mahoganybuttstrings;optimistix",
        forked_from_public_notebook="FALSE",
        academic_papers_cited="",
        uses_canonized_technique="HC;Ridge-as-stacker;target-encoding-stack;DAE-as-base-encoder;original-as-columns",
        author_surprise_flag="FALSE", author_recurring_in_set="TRUE", author_appearance_count=8,
        notable_commenters="cdeotte;tilii7",
        winner_unique_edge="Validation discipline headline: chose 0.9557801-CV submission over 0.955865 because CV-LB degraded past 0.95578; would have been 3rd if greedy; 150 OOFs + Optuna selects ~10%",
        reeval_doc="analysis/writeup-reevaluation/s6e2_rank1.md",
    ),
    # s6e3 - cdeotte - mixed (KGMON Playbook: 39 forks + own monster + 4-level stack)
    dict(
        competition_ref="playground-series-s6e3", finish_rank=1, author="cdeotte",
        paradigm="ensemble-stacking", lookup_exploit_subtype="none", lookup_material_present="TRUE",
        origination_score=2,
        cited_community_members="angelosmar1;azzamradman;badalkrsharma;blamerx;furqonaryadana;greysky;include4eto;johnnyhyland;lightningv08;mahoganybuttstrings;masayakawamata;mikhailnaumov;nalgirayergn;omidbaghchehsaraei;ozermehmet;rawashishsin;siukeitin;tsunamazda;yekenot;yunsuxiaozi;arunklenin",
        forked_from_public_notebook="FALSE",
        academic_papers_cited="",
        uses_canonized_technique="HC;brute-force-FE;Ridge-as-stacker;RAPIDS-XGB;target-encoding-stack;original-as-columns;pseudo-labeling;DAE-as-base-encoder",
        author_surprise_flag="FALSE", author_recurring_in_set="TRUE", author_appearance_count=8,
        notable_commenters="mahog;masayakawamata",
        winner_unique_edge="KGMON Playbook at industrial scale: 850 candidate models from 39 public notebook ports + LLM 'AI Council' (GPT5.4+Gemini3.1+Claude4.6) generating 600K LOC; 4-level stack; snap+KDTree on IBM original",
        reeval_doc="analysis/writeup-reevaluation/s6e3_rank1.md",
    ),
    # s6e4 - kirill0212 - community-template-tweak (90% public OOFs + meta-architecture)
    dict(
        competition_ref="playground-series-s6e4", finish_rank=1, author="kirill0212 (cstdy)",
        paradigm="community-template-tweak", lookup_exploit_subtype="none", lookup_material_present="FALSE",
        origination_score=1,
        cited_community_members="wguesdon;mahoganybuttstrings;utaazu;yunsuxiaozi;include4eto;yekenot;mikhail_naumov;maria_nadeem;lucas_moraes;rohit;ashish;gnn",
        forked_from_public_notebook="TRUE",
        academic_papers_cited="",
        uses_canonized_technique="Ridge-as-stacker;target-encoding-stack",
        author_surprise_flag="FALSE", author_recurring_in_set="FALSE", author_appearance_count=1,
        notable_commenters="mahog",
        winner_unique_edge="~90% of OOFs from other Kagglers (wguesdon 30/49% + 24 others); win = meta-architecture: two-binary-classifier reframe of 3-class (Low-vs-Rest + Med-vs-High) + logit LR blend + threshold search",
        reeval_doc="analysis/writeup-reevaluation/s6e4_rank1.md",
    ),
]


def main():
    if not XLSX_PATH.exists():
        print(f"ERROR: {XLSX_PATH} not found")
        sys.exit(1)

    wb = openpyxl.load_workbook(XLSX_PATH)

    # Remove existing sheet if present (idempotent rebuild)
    if SHEET_NAME in wb.sheetnames:
        del wb[SHEET_NAME]

    ws = wb.create_sheet(SHEET_NAME)

    # Header row
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    for col_idx, hdr in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=hdr)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # Data rows
    for row_idx, row in enumerate(ROWS, start=2):
        # auto-compute n_cited_members from cited_community_members
        cited = row.get("cited_community_members", "")
        n_cited = 0
        if cited.strip():
            n_cited = len([s for s in cited.split(";") if s.strip()])

        # winner_unique_edge length check
        edge = row.get("winner_unique_edge", "")
        if len(edge) > 200:
            print(f"WARN row {row_idx} ({row['competition_ref']}): unique_edge={len(edge)} chars > 200")

        values = [
            row.get("competition_ref", ""),
            row.get("finish_rank", ""),
            row.get("author", ""),
            row.get("paradigm", ""),
            row.get("lookup_exploit_subtype", ""),
            row.get("lookup_material_present", ""),
            row.get("origination_score", ""),
            cited,
            n_cited,
            row.get("forked_from_public_notebook", ""),
            row.get("academic_papers_cited", ""),
            row.get("uses_canonized_technique", ""),
            row.get("author_surprise_flag", ""),
            row.get("author_recurring_in_set", ""),
            row.get("author_appearance_count", ""),
            row.get("notable_commenters", ""),
            edge,
            row.get("reeval_doc", ""),
        ]
        for col_idx, v in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=v)
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # Column widths (rough heuristic)
    widths = {
        "A": 38, "B": 6, "C": 30, "D": 24, "E": 28, "F": 8, "G": 6,
        "H": 50, "I": 6, "J": 8, "K": 24, "L": 38, "M": 8, "N": 8,
        "O": 6, "P": 36, "Q": 60, "R": 50,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    ws.freeze_panes = "A2"

    # Update Codebook sheet
    cb = wb["Codebook"]
    # Find next empty row
    next_row = cb.max_row + 1
    # Spacer header
    spacer_cell = cb.cell(row=next_row, column=1, value="--- Pass 2 (Paradigm & Attribution sheet) ---")
    spacer_cell.font = Font(bold=True, italic=True)
    next_row += 1
    for field, allowed in CODEBOOK_ADDITIONS:
        cb.cell(row=next_row, column=1, value=field)
        c = cb.cell(row=next_row, column=2, value=allowed)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        next_row += 1

    wb.save(XLSX_PATH)
    print(f"OK: wrote sheet '{SHEET_NAME}' with {len(ROWS)} rows, updated Codebook (+{len(CODEBOOK_ADDITIONS)} entries)")
    print(f"Rows by paradigm:")
    from collections import Counter
    para_counts = Counter(r["paradigm"] for r in ROWS)
    for p, c in sorted(para_counts.items(), key=lambda x: -x[1]):
        print(f"  {p:30s} {c}")
    print(f"\nLookup-exploit subtypes (where present):")
    sub_counts = Counter(r["lookup_exploit_subtype"] for r in ROWS if r["lookup_exploit_subtype"] != "none")
    for s, c in sorted(sub_counts.items(), key=lambda x: -x[1]):
        print(f"  {s:30s} {c}")
    print(f"\nlookup_material_present TRUE: {sum(1 for r in ROWS if r['lookup_material_present'] == 'TRUE')} / {len(ROWS)}")
    print(f"Origination score distribution:")
    orig_counts = Counter(r["origination_score"] for r in ROWS)
    for s, c in sorted(orig_counts.items()):
        print(f"  {s}: {c}")
    print(f"\nSurprise wins: {sum(1 for r in ROWS if r['author_surprise_flag'] == 'TRUE')}")
    print(f"Forked from public notebook: {sum(1 for r in ROWS if r['forked_from_public_notebook'] == 'TRUE')}")
    print(f"Academic paper cited: {sum(1 for r in ROWS if r['academic_papers_cited'])}")


if __name__ == "__main__":
    main()

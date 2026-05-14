"""
Normalize synonym variants across coded fields in kaggle_meta_analysis.xlsx.
Prints every change made for review.
"""
import openpyxl
from pathlib import Path

XL_PATH = Path(__file__).parent.parent / "data" / "kaggle_meta_analysis.xlsx"

wb = openpyxl.load_workbook(XL_PATH)
ws = wb["Competition Data"]
headers = [c.value for c in ws[1]]

def col(name):
    return headers.index(name) + 1

changes = []

def normalize(r, field, new_val):
    c = col(field)
    old = ws.cell(r, c).value
    if str(old) != str(new_val):
        ws.cell(r, c).value = new_val
        slug = ws.cell(r, col("competition_ref")).value
        changes.append(f"  {slug:<50} {field}: {old!r} -> {new_val!r}")

# ---------------------------------------------------------------
# Rules — applied row by row
# ---------------------------------------------------------------
for r in range(2, ws.max_row + 1):
    slug = ws.cell(r, col("competition_ref")).value
    if not slug:
        continue

    def get(field):
        return ws.cell(r, col(field)).value

    # --- has_categorical: normalize to uppercase string TRUE/FALSE ---
    v = str(get("has_categorical")).strip()
    if v.lower() == "true":
        normalize(r, "has_categorical", "TRUE")
    elif v.lower() == "false":
        normalize(r, "has_categorical", "FALSE")

    # --- primary_model: consolidate specific model names to types ---
    v = str(get("primary_model") or "").strip().lower()
    if v in ("xgboost", "lightgbm", "catboost", "gbm"):
        normalize(r, "primary_model", "gbm")
    elif v == "nn":
        normalize(r, "primary_model", "neural_network")
    elif v == "ensemble_mixed":
        normalize(r, "primary_model", "ensemble")

    # --- encoding_strategy: normalize synonyms, standardize separator to "; " ---
    v = str(get("encoding_strategy") or "").strip()
    # split on comma or semicolon
    parts = [p.strip() for p in v.replace(";", ",").split(",") if p.strip()]
    mapped = []
    for p in parts:
        pl = p.lower()
        if pl in ("target", "target_encoding"):
            mapped.append("target_encoding")
        elif pl in ("ohe", "one_hot", "one_hot_encoding"):
            mapped.append("one_hot_encoding")
        elif pl in ("label", "label_encoding"):
            mapped.append("label_encoding")
        elif pl == "ordinal":
            mapped.append("ordinal_encoding")
        elif pl == "catboost_encoding":
            mapped.append("target_encoding")  # CatBoost encoding is a form of TE
        elif pl == "count_encoding":
            mapped.append("count_encoding")
        else:
            mapped.append(p)  # keep as-is (none, not_described, etc.)
    # deduplicate preserving order
    seen = set()
    deduped = []
    for m in mapped:
        if m not in seen:
            seen.add(m)
            deduped.append(m)
    new_v = "; ".join(deduped)
    normalize(r, "encoding_strategy", new_v)

    # --- scaling: standard_scaler -> standard; fix "not described" ---
    v = str(get("scaling") or "").strip()
    if v == "standard_scaler":
        normalize(r, "scaling", "standard")
    elif v == "not described":
        normalize(r, "scaling", "not_described")

    # --- ensemble_method: normalize blend variants; standardize separator ---
    v = str(get("ensemble_method") or "").strip()
    parts = [p.strip() for p in v.replace(";", ",").split(",") if p.strip()]
    mapped = []
    for p in parts:
        pl = p.lower()
        if pl in ("blend", "blending"):
            mapped.append("mean_blend")
        elif pl in ("weighted_blending"):
            mapped.append("weighted_blend")
        elif pl == "ridge_stacking":
            mapped.append("stacking")
        elif pl == "ridge_ensemble":
            mapped.append("stacking")
        elif pl == "nn_ensemble":
            mapped.append("stacking")
        elif pl == "catboost_residual_baseline":
            mapped.append("stacking")
        else:
            mapped.append(p)
    seen = set()
    deduped = []
    for m in mapped:
        if m not in seen:
            seen.add(m)
            deduped.append(m)
    new_v = "; ".join(deduped)
    normalize(r, "ensemble_method", new_v)

    # --- original_data_usage: normalize True/False/yes/not_used ---
    v = str(get("original_data_usage") or "").strip()
    if v in ("True", "yes"):
        normalize(r, "original_data_usage", "yes")
    elif v in ("False", "not_used"):
        normalize(r, "original_data_usage", "none")

    # --- distribution_shift: normalize casing and "not described" / "low" ---
    v = str(get("distribution_shift") or "").strip()
    if v == "not described":
        normalize(r, "distribution_shift", "not_described")
    elif v.lower() == "true":
        normalize(r, "distribution_shift", "TRUE")
    elif v.lower() == "false":
        normalize(r, "distribution_shift", "FALSE")
    elif v.lower() == "low":
        normalize(r, "distribution_shift", "TRUE")  # some shift = TRUE

    # --- outlier_treatment: fix stray "not described" ---
    v = str(get("outlier_treatment") or "").strip()
    if v == "not described":
        normalize(r, "outlier_treatment", "not_described")

    # --- rare_class_handling: fix stray "not described" ---
    v = str(get("rare_class_handling") or "").strip()
    if v == "not described":
        normalize(r, "rare_class_handling", "not_described")

    # --- missing_data_strategy: fix stray "not described" ---
    v = str(get("missing_data_strategy") or "").strip()
    if v == "not described":
        normalize(r, "missing_data_strategy", "not_described")

wb.save(XL_PATH)
print(f"{len(changes)} changes made:\n")
for c in changes:
    print(c)
print("\nDone.")

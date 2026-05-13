"""One-off: correct finish_rank for s5e3 from 1 to 2 (1st place had no writeup or notebook)."""
import openpyxl
from pathlib import Path

ROOT = Path(__file__).parent.parent
XL_PATH = ROOT / "data" / "kaggle_meta_analysis.xlsx"

wb = openpyxl.load_workbook(XL_PATH)
ws = wb["Competition Data"]
headers = [c.value for c in ws[1]]
ref_col  = headers.index("competition_ref") + 1
rank_col = headers.index("finish_rank") + 1

for r in range(2, ws.max_row + 1):
    if str(ws.cell(r, ref_col).value) == "playground-series-s5e3":
        old = ws.cell(r, rank_col).value
        ws.cell(r, rank_col).value = 2
        print(f"s5e3: finish_rank {old} -> 2")
        break

wb.save(XL_PATH)

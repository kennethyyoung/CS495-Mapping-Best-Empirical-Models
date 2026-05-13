"""Fill the metric column using the Kaggle API competition detail endpoint."""
import openpyxl
from pathlib import Path
from kaggle import KaggleApi
import json

ROOT    = Path(__file__).parent.parent
XL_PATH = ROOT / "data" / "kaggle_meta_analysis.xlsx"

api = KaggleApi()
api.authenticate()

wb = openpyxl.load_workbook(XL_PATH)
ws = wb["Competition Data"]
headers = [c.value for c in ws[1]]

def col(name):
    return headers.index(name) + 1

BLANK = (None, "", "not_described")

for r in range(2, ws.max_row + 1):
    slug = ws.cell(r, col("competition_ref")).value
    if not slug:
        continue
    if ws.cell(r, col("metric")).value not in BLANK:
        print(f"SKIP  {slug}")
        continue

    print(f"FETCH {slug} ...", end=" ", flush=True)
    try:
        resp = api.competitions_list(search=slug)
        competitions = resp.competitions  # ApiListCompetitionsResponse
        match = next((c for c in competitions if slug in str(c.ref)), None)
        if match:
            metric = str(match.evaluation_metric) if match.evaluation_metric else "not_described"
        else:
            metric = "not_described"
        ws.cell(r, col("metric")).value = metric
        print(metric)
    except Exception as e:
        print(f"ERROR: {e}")
        ws.cell(r, col("metric")).value = "not_described"

wb.save(XL_PATH)
print("\nDone.")

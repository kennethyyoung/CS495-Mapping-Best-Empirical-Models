"""Split original_data_usage into external_original_available + external_original_use_mode.

The old column conflates two distinct questions ("is an external public
original dataset available?" and "if so, how is it used in the solution?").
Re-eval surfaced cases where this conflation hides meaningful decisions:
- cdeotte's s3e5 chose NOT to use available original (CV-driven decision)
  vs s4e12 had no available original; old column codes both as 'none'.
- s5e10 (Tilii) was coded 'yes' but the 'original' use was internal-data
  augmentation, not external dataset (corrected to 'none' in prior commit).

Old column is preserved in-place with a deprecation note in the Codebook.
Two new columns are added alongside.

Mapping derived per-entry from re-eval docs (see source field).
"""

import sys
from datetime import date
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

sys.stdout.reconfigure(encoding='utf-8')

XLSX_PATH = Path("data/kaggle_meta_analysis.xlsx")

# Mapping: (competition_ref, finish_rank) -> (available, use_mode, source_doc, notes)
#   available: TRUE / FALSE / unknown
#   use_mode: rows-only | columns-only | both | features-derived | lookup |
#             available-not-used | unavailable | unknown
MAPPING = {
    ("icr-identify-age-related-conditions", 1): ("FALSE", "unavailable", "icr_rank1.md", ""),
    ("playground-series-s4e7", 1): ("TRUE", "rows-only", "s4e7_rank1.md", "Public original + target reversal trick on duplicates."),
    ("playground-series-s5e3", 2): ("TRUE", "both", "s5e3_rank2.md", "cdeotte's deliberate dual-mode use (rows for XGB+TabPFN, columns for SVC)."),
    ("playground-series-s5e5", 1): ("TRUE", "rows-only", "s5e5_rank1.md", "Minimally used (concat only); not column-exploited."),
    ("playground-series-s5e12", 1): ("TRUE", "both", "s5e12_rank1.md", "Orig_as_columns base model + concat for others."),
    ("playground-series-s5e10", 1): ("FALSE", "unavailable", "s5e10_rank1.md", "No external original; Lasso reverse-engineered the generator formula from training data (corrected from 'yes' on 2026-05-27)."),
    ("playground-series-s4e10", 2): ("TRUE", "rows-only", "s4e10_rank2.md", "Public original concat_rows."),
    ("playground-series-s5e11", 1): ("TRUE", "rows-only", "s5e11_rank1.md", "Public loan dataset concatenated for training rows."),
    ("playground-series-s4e1", 3): ("TRUE", "rows-only", "s4e1_rank3.md", "Public bank-churn dataset; CatBoost-encoder-order-sensitivity discovery."),
    ("playground-series-s5e2", 1): ("TRUE", "columns-only", "s5e2_rank1.md", "cdeotte 'manufacture suggested retail price' columns-only feature."),
    ("playground-series-s5e8", 2): ("TRUE", "both", "s5e8_rank2.md", "Concat + TE on bigrams using BOTH competition + original targets."),
    ("playground-series-s5e4", 2): ("unknown", "unknown", "s5e4_rank2.md", "Writeup doesn't explicitly cite external original use."),
    ("playground-series-s4e9", 1): ("TRUE", "rows-only", "s4e9_rank1.md", "Public original car-price dataset concat."),
    ("playground-series-s4e5", 1): ("TRUE", "available-not-used", "s4e5_rank1.md", "Original existed but 'completely different from train dataset' — author used for EDA only."),
    ("playground-series-s4e11", 1): ("TRUE", "rows-only", "s4e11_rank1.md", "Public original concat."),
    ("playground-series-s5e6", 1): ("TRUE", "both", "s5e6_rank1.md", "Dual-mode: concat_rows for some models + columns (TE against original) for the 2268-feature monster."),
    ("playground-series-s4e8", 1): ("TRUE", "both", "s4e8_rank1.md", "Concat + siukeitin's 'exact solution to original' feature (features-derived from original target)."),
    ("playground-series-s4e12", 1): ("FALSE", "unavailable", "s4e12_rank1.md", "First cdeotte writeup without public original."),
    ("playground-series-s4e3", 2): ("TRUE", "rows-only", "s4e3_rank2.md", "Public UCI steel-plates dataset concat."),
    ("playground-series-s3e24", 3): ("TRUE", "rows-only", "s3e24_rank3.md", "Public original concat."),
    ("playground-series-s3e23", 2): ("FALSE", "unavailable", "s3e23_rank2.md", "No public original used."),
    ("playground-series-s3e14", 1): ("TRUE", "lookup", "s3e14_rank1.md", "Lookup exploit: match test rows to original by (fruitset, fruitmass) keys; primary winning mechanism."),
    ("playground-series-s3e26", 2): ("unknown", "unknown", "s3e26_rank2.md", "Author didn't disclose external-original use; not_described in Pass 1."),
    ("playground-series-s3e16", 3): ("TRUE", "rows-only", "s3e16_rank3.md", "Public original concat + generator-driven additional synthetic data."),
    ("playground-series-s3e17", 3): ("unknown", "unknown", "s3e17_rank3.md", "Writeup minimal (~30 lines); external-original use not described."),
    ("playground-series-s3e11", 1): ("TRUE", "rows-only", "s3e11_rank1.md", "Public original media-campaign dataset concat."),
    ("playground-series-s3e13", 1): ("FALSE", "unavailable", "s3e13_rank1.md", "No public original used."),
    ("playground-series-s3e5", 1): ("TRUE", "available-not-used", "s3e5_rank1.md", "Original existed but author chose NOT to use (overfit local CV)."),
    ("playground-series-s3e10", 1): ("FALSE", "unavailable", "s3e10_rank1.md", "No public original used."),
    ("playground-series-s3e9", 1): ("FALSE", "unavailable", "s3e9_rank1.md", "No public original used."),
    ("playground-series-s3e8", 2): ("TRUE", "rows-only", "s3e8_rank2.md", "Public gemstone dataset concat; lookup-exploit TESTED but rejected (overfit public LB)."),
    ("playground-series-s3e6", 3): ("TRUE", "rows-only", "s3e6_rank3.md", "Public Paris housing dataset concat."),
    ("playground-series-s3e1", 1): ("TRUE", "rows-only", "s3e1_rank1.md", "Public California Housing dataset concat."),
    ("playground-series-s3e7", 1): ("TRUE", "both", "s3e7_rank1.md", "Public original concat + adversarial-validation filter; inversion-lookup exploit."),
    ("playground-series-s3e3", 1): ("TRUE", "rows-only", "s3e3_rank1.md", "Public IBM HR dataset concat."),
    ("playground-series-s5e7", 2): ("TRUE", "lookup", "s5e7_rank2.md", "match_p trick: exact-row-join with public original on 7-feature tuple."),
    ("playground-series-s4e4", 1): ("TRUE", "rows-only", "s4e4_rank1.md", "Public UCI Abalone dataset; train-on-original-only-no-validate scheme for distribution-shift handling."),
    ("playground-series-s3e4", 3): ("FALSE", "unavailable", "s3e4_rank3.md", "No external original used (PCA-anonymized features make exploit-original infeasible)."),
    ("tabular-playground-series-feb-2022", 1): ("FALSE", "unavailable", "tps_feb_2022_rank1.md", "No public original; exploit was source-code-level reverse-engineering of np.random."),
    ("tabular-playground-series-may-2022", 1): ("FALSE", "unavailable", "tps_may_2022_rank1.md", "No public original."),
    ("tabular-playground-series-jun-2022", 1): ("FALSE", "unavailable", "tps_jun_2022_rank1.md", "No public original."),
    ("playground-series-s6e1", 1): ("FALSE", "unavailable", "s6e1_rank1.md", "No external dataset used; first of s6 quartet without one."),
    ("playground-series-s6e2", 1): ("TRUE", "both", "s6e2_rank1.md", "Public heart-disease original + target stats (mean/smoothed/WoE/entropy) as external TE features."),
    ("playground-series-s6e3", 1): ("TRUE", "both", "s6e3_rank1.md", "IBM Telco Customer Churn (7,032 rows) + snap features + cKDTree nearest-neighbor lookup + PCA/GRP projection fitted on original."),
    ("playground-series-s6e4", 1): ("TRUE", "rows-only", "s6e4_rank1.md", "External miadul/irrigation-water-requirement-prediction-dataset concatenated."),
}

CODEBOOK_ENTRIES = [
    ("--- Pass 1 split (added 2026-05-27): original_data_usage ---", ""),
    ("original_data_usage (DEPRECATED)",
     "DEPRECATED 2026-05-27. Conflated availability and use mode. Use external_original_available + external_original_use_mode instead. Old values retained for audit trail; new analyses should reference the new columns."),
    ("external_original_available",
     "TRUE | FALSE | unknown. TRUE if a public external original dataset exists for this competition (regardless of whether the winner used it). 'unknown' when writeup is silent on this."),
    ("external_original_use_mode",
     "rows-only | columns-only | both | features-derived | lookup | available-not-used | unavailable | unknown. How the external original is used: rows-only = concatenated as training rows; columns-only = original-as-columns (cdeotte's merge_columns); both = concat + columns; features-derived = aggregated stats from original used as features (e.g. target-mean TE on original); lookup = direct test-to-original key-match or KDTree (the lookup-exploit family); available-not-used = original exists but author deliberately did not use (cv-driven decision); unavailable = no external original exists; unknown = writeup silent."),
]


def main():
    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb["Competition Data"]
    headers = [c.value for c in ws[1]]

    def col(name):
        return headers.index(name) + 1

    # Add the two new columns after original_data_usage (or at end if not present)
    if "external_original_available" not in headers:
        new_col = ws.max_column + 1
        ws.cell(row=1, column=new_col, value="external_original_available").font = Font(bold=True)
        headers.append("external_original_available")
        print(f"added external_original_available at col {new_col}")
    if "external_original_use_mode" not in headers:
        new_col = ws.max_column + 1
        ws.cell(row=1, column=new_col, value="external_original_use_mode").font = Font(bold=True)
        headers.append("external_original_use_mode")
        print(f"added external_original_use_mode at col {new_col}")

    avail_c = col("external_original_available")
    mode_c = col("external_original_use_mode")
    ref_c = col("competition_ref")
    rank_c = col("finish_rank")

    # Apply mapping
    changes = []
    filled = 0
    missing = []
    for r in range(2, ws.max_row + 1):
        comp_ref = ws.cell(row=r, column=ref_c).value
        rank = ws.cell(row=r, column=rank_c).value
        key = (comp_ref, rank)
        if key not in MAPPING:
            missing.append(key)
            continue
        avail, mode, src, notes = MAPPING[key]
        old_avail = ws.cell(row=r, column=avail_c).value
        old_mode = ws.cell(row=r, column=mode_c).value
        if old_avail != avail:
            ws.cell(row=r, column=avail_c, value=avail)
            changes.append((comp_ref, rank, "external_original_available", old_avail, avail, src, notes))
        if old_mode != mode:
            ws.cell(row=r, column=mode_c, value=mode)
            changes.append((comp_ref, rank, "external_original_use_mode", old_mode, mode, src, notes))
        filled += 1

    print(f"populated {filled}/{ws.max_row - 1} rows")
    if missing:
        print(f"WARNING — {len(missing)} rows had no mapping: {missing[:5]}")

    # Append codebook entries (skip if already present)
    cb = wb["Codebook"]
    existing = set()
    for r in range(2, cb.max_row + 1):
        v = cb.cell(row=r, column=1).value
        if v:
            existing.add(v.strip())
    next_row = cb.max_row + 1
    added = 0
    for field, allowed in CODEBOOK_ENTRIES:
        if field.strip() in existing:
            continue
        cb.cell(row=next_row, column=1, value=field)
        c = cb.cell(row=next_row, column=2, value=allowed)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        if field.startswith("---"):
            cb.cell(row=next_row, column=1).font = Font(bold=True, italic=True)
        next_row += 1
        added += 1
    print(f"added {added} codebook entries")

    # Append to Corrections Log
    if "Corrections Log" in wb.sheetnames:
        cl = wb["Corrections Log"]
        next_log_row = cl.max_row + 1
        for comp_ref, rank, field, old, new, src, notes in changes:
            severity = "medium"
            cl.cell(row=next_log_row, column=1, value=str(date.today()))
            cl.cell(row=next_log_row, column=2, value=comp_ref)
            cl.cell(row=next_log_row, column=3, value=rank)
            cl.cell(row=next_log_row, column=4, value=field)
            cl.cell(row=next_log_row, column=5, value=str(old) if old is not None else "")
            cl.cell(row=next_log_row, column=6, value=str(new))
            cl.cell(row=next_log_row, column=7, value=severity)
            cl.cell(row=next_log_row, column=8, value=f"Backfilled from original_data_usage split. {notes}".strip())
            cl.cell(row=next_log_row, column=9, value=f"analysis/writeup-reevaluation/{src}")
            fill = PatternFill(start_color="FFF4E6", end_color="FFF4E6", fill_type="solid")
            for c_idx in range(1, 10):
                cell = cl.cell(row=next_log_row, column=c_idx)
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                cell.fill = fill
            next_log_row += 1
        print(f"appended {len(changes)} entries to Corrections Log")

    wb.save(XLSX_PATH)
    print("\nOK saved.")
    print("\nDistribution of external_original_available:")
    from collections import Counter
    avail_counts = Counter(m[0] for m in MAPPING.values())
    for k, v in sorted(avail_counts.items()):
        print(f"  {k}: {v}")
    print("\nDistribution of external_original_use_mode:")
    mode_counts = Counter(m[1] for m in MAPPING.values())
    for k, v in sorted(mode_counts.items(), key=lambda x: -x[1]):
        print(f"  {k}: {v}")


if __name__ == "__main__":
    main()

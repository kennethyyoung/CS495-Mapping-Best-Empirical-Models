import os
os.chdir(r"C:\Users\Nebi\Documents\capstone\CS495-Mapping-Best-Empirical-Models")
import openpyxl
wb = openpyxl.load_workbook("data/kaggle_candidates_v2.xlsx")
ws = wb.active
headers = [c.value for c in ws[1]]
print("Columns:", headers)
print("Total rows:", ws.max_row - 1)
# Show one sample row
row2 = {headers[i]: ws.cell(2, i+1).value for i in range(len(headers))}
print("\nSample row:")
for k, v in row2.items():
    print(f"  {k}: {v}")

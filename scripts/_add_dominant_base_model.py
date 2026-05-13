"""
Add dominant_base_model column to kaggle_meta_analysis.xlsx.
Derives value from best_single_model and models_used using keyword matching.
Flags entries that need manual review.
"""
import openpyxl
from pathlib import Path

XL_PATH = Path(__file__).parent.parent / "data" / "kaggle_meta_analysis.xlsx"

GBM_KEYWORDS    = ["xgboost", "xgb", "lightgbm", "lgbm", "catboost", "gbm", "gbdt",
                   "gradient boost", "histgradient", "ydf", "random forest", "rf"]
NN_KEYWORDS     = ["realmlp", "neural", " nn", "mlp", "tabnet", "transformer",
                   "tabm", "tabicl", "ft-transformer", "tabtransformer", "gandalf",
                   "resnet", "deepfm", "ffm", "saint", "fastai", "pytorch", "keras",
                   "embedding", "liquid", "trompt", "gnn", "daenet", "dae", "vsn",
                   "tabpfn", "nam", "dcn", "snn", "ifan"]
LINEAR_KEYWORDS = ["ridge", "logistic", "lasso", "linear", "logreg", "lr"]

def classify(text):
    if not text:
        return None
    t = str(text).lower()
    scores = {"gbm": 0, "neural_network": 0, "linear": 0}
    for kw in GBM_KEYWORDS:
        if kw in t:
            scores["gbm"] += 1
    for kw in NN_KEYWORDS:
        if kw in t:
            scores["neural_network"] += 1
    for kw in LINEAR_KEYWORDS:
        if kw in t:
            scores["linear"] += 1
    if max(scores.values()) == 0:
        return None
    return max(scores, key=scores.get)

wb = openpyxl.load_workbook(XL_PATH)
ws = wb["Competition Data"]
headers = [c.value for c in ws[1]]

# Add column if not present
if "dominant_base_model" not in headers:
    new_col = len(headers) + 1
    ws.cell(1, new_col).value = "dominant_base_model"
    headers.append("dominant_base_model")
    print(f"Added dominant_base_model at column {new_col}")
else:
    new_col = headers.index("dominant_base_model") + 1
    print("dominant_base_model column already exists")

ref_col  = headers.index("competition_ref") + 1
bsm_col  = headers.index("best_single_model") + 1
mu_col   = headers.index("models_used") + 1
pm_col   = headers.index("primary_model") + 1

BLANK = (None, "", "not_described", "not described")

needs_review = []

for r in range(2, ws.max_row + 1):
    slug      = ws.cell(r, ref_col).value
    if not slug:
        continue

    existing  = ws.cell(r, new_col).value
    if existing not in BLANK:
        print(f"SKIP  {slug} (already: {existing})")
        continue

    bsm       = ws.cell(r, bsm_col).value
    mu        = ws.cell(r, mu_col).value
    primary   = ws.cell(r, pm_col).value

    # Priority: best_single_model > models_used > primary_model
    result = classify(bsm) or classify(mu)

    # If primary_model is already a specific type (not "ensemble"), use it as fallback
    if result is None and primary not in (None, "", "ensemble", "ensemble_mixed", "other"):
        p = str(primary).lower()
        if any(k in p for k in ["xgb", "lgbm", "catboost", "gbm", "lightgbm"]):
            result = "gbm"
        elif any(k in p for k in ["neural", "nn", "mlp"]):
            result = "neural_network"
        elif any(k in p for k in ["ridge", "logistic", "linear"]):
            result = "linear"

    if result:
        ws.cell(r, new_col).value = result
        print(f"SET   {slug:<55} -> {result}  (from bsm={bsm!r})")
    else:
        needs_review.append((slug, bsm, mu, primary))
        print(f"FLAG  {slug:<55}  bsm={bsm!r}, primary={primary!r}")

wb.save(XL_PATH)
print(f"\nDone. {len(needs_review)} entries need manual review:")
for slug, bsm, mu, pm in needs_review:
    print(f"  {slug}: bsm={bsm!r}, primary_model={pm!r}")

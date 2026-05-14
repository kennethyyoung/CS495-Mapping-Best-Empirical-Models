import os
os.chdir(r"C:\Users\Nebi\Documents\capstone\CS495-Mapping-Best-Empirical-Models")
import openpyxl
import pandas as pd
from pathlib import Path

RAW = Path("data/raw")
wb = openpyxl.load_workbook("data/kaggle_meta_analysis.xlsx")
ws = wb["Competition Data"]
headers = [c.value for c in ws[1]]
ref_col = headers.index("competition_ref") + 1
nf_col = headers.index("n_features") + 1
pm_col = headers.index("pct_missing") + 1

for r in range(2, ws.max_row + 1):
    slug = ws.cell(r, ref_col).value
    nf = ws.cell(r, nf_col).value
    pm = ws.cell(r, pm_col).value
    train = RAW / str(slug) / "train.csv"
    has_file = train.exists()
    print(f"{slug:45s}  n_features={str(nf):12s}  pct_missing={str(pm):12s}  train.csv={'yes' if has_file else 'MISSING'}")

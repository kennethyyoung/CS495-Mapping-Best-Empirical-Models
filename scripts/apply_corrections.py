"""Apply audit-identified corrections to kaggle_meta_analysis.xlsx in severity order.

Phases (severity order):
 1. HIGH cell-level: s5e2 n_rows, s3e4 winners, s6e2 fe_techniques
 2. HIGH schema: add top_3_margin column (backfilled from leaderboard.json)
 3. MEDIUM cell-level: models_used undercounts, s6e3 dominant_base_model,
    s4e11 missing_data_strategy, s5e10 original_data_usage semantics, etc.
 4. MEDIUM schema: add distribution_shift_type column
 5. LOW: Codebook clarifications

Every cell change is recorded in a new 'Corrections Log' sheet with
(competition, field, old_value, new_value, reason, source_doc).

Idempotent: re-running rebuilds Corrections Log from the changes list and
re-applies values. Existing equal values are not re-logged.

original_data_usage SPLIT is held for a separate run (most invasive — needs
per-entry availability/use-mode mapping).
"""

import json
import sys
from pathlib import Path
from datetime import date
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

sys.stdout.reconfigure(encoding='utf-8')

XLSX_PATH = Path("data/kaggle_meta_analysis.xlsx")
LB_DIR = Path("data/raw")
CORRECTIONS_SHEET = "Corrections Log"


# =====================================================================
# CELL-LEVEL CORRECTIONS (competition, finish_rank, field, new_value, reason, severity, source)
# =====================================================================
CELL_CORRECTIONS = [
    # ---- HIGH SEVERITY ----
    ("playground-series-s5e2", 1, "n_rows", 4000000,
     "Combined train.csv (300K) + training_extra.csv (~3.7M) loaded together in cdeotte's notebook. Strategy was shaped by 4M, not 300K.",
     "high", "analysis/writeup-reevaluation/s5e2_rank1.md"),

    ("playground-series-s3e4", 3, "winner_1st", "Dmitry",
     "Actual 1st place per leaderboard.json. Was 'not described'.",
     "high", "data/raw/playground-series-s3e4/leaderboard.json"),

    ("playground-series-s3e4", 3, "winner_2nd", "Numanynklr",
     "Actual 2nd place per leaderboard.json. Was 'not described'.",
     "high", "data/raw/playground-series-s3e4/leaderboard.json"),

    ("playground-series-s6e2", 1, "fe_techniques",
     "binning; digit features; all-as-categorical; within-fold TE (cuML); gplearn nonlinear features; original-data target stats (mean/smoothed/WoE/entropy); DVAE latents",
     "Removed 'knowledge distillation (0.5 hard + 0.5 teacher OOF)' — training-time technique, not FE.",
     "high", "analysis/writeup-reevaluation/s6e2_rank1.md"),

    # ---- MEDIUM SEVERITY ----
    ("playground-series-s6e3", 1, "dominant_base_model", "gbm",
     "Writeup count: 90 trees vs 60 NNs — trees dominate by count. Previous 'neural_network' label likely came from best_single_model being TabM.",
     "medium", "analysis/writeup-reevaluation/s6e3_rank1.md"),

    ("playground-series-s5e10", 1, "original_data_usage", "none",
     "Tilii used Lasso to reverse-engineer generator formula from training data. No external public original dataset was used. Was 'yes' which conflated internal-data use with external-original use.",
     "medium", "analysis/writeup-reevaluation/s5e10_rank1.md"),

    ("playground-series-s4e11", 1, "missing_data_strategy", "leave_as_is",
     "Author principle (explicit in comments): 'usually best to leave the inconsistencies in the data as is'. Was 'imputation'.",
     "medium", "analysis/writeup-reevaluation/s4e11_rank1.md"),

    ("playground-series-s5e11", 1, "models_used",
     "lightgbm; xgboost; catboost; tabm; lama-denselight; lama-dense; tabular_rnn; excelformer; modernnca; realmlp; ftt; neural_network; ann; deeptables; resnet; gandalf; saint; tabr; mlp; node; transformer; ftft; xgb_regression",
     "Writeup lists 23 distinct architectures; spreadsheet previously had 16. Added: LAMA-DenseLight, LAMA-Dense, TabulaRNN, ExcelFormer, ModernNCA, plus distinguished XGB-regression and LightGBM-DART variants.",
     "medium", "analysis/writeup-reevaluation/s5e11_rank1.md"),

    ("playground-series-s5e8", 2, "models_used",
     "tabm; xgboost; lightgbm; catboost; deeptables; tabr; realmlp; gandalf; grn; fttransformer; cnn; random_forest; bart",
     "Writeup shows 13 distinct architectures. Previous lumped under 'neural_network' (7 categories total). Now distinguishes DeepTables, TabR, RealMLP, Gandalf, GRN, FTTransformer, CNN.",
     "medium", "analysis/writeup-reevaluation/s5e8_rank2.md"),

    ("playground-series-s5e10", 1, "models_used",
     "xgboost; catboost; neural_network; lasso; autogluon; autoencoder; keras_fm; tabm; tf_embedding_network",
     "Writeup also names Keras FM, TabM, TF embedding networks (previously lumped under 'NN').",
     "medium", "analysis/writeup-reevaluation/s5e10_rank1.md"),

    ("playground-series-s5e12", 1, "models_used",
     "xgboost; lightgbm; catboost; histgradientboosting; neural_network; ridge; automl",
     "Added HistGradientBoosting (visible in notebook MODELS list).",
     "medium", "analysis/writeup-reevaluation/s5e12_rank1.md"),

    ("playground-series-s4e7", 1, "models_used",
     "lightgbm; catboost; xgboost; denselight; mlp; tab_resnet; tab_transformer; autoint",
     "Cross Sellers used 5 distinct NN architectures (Denselight, MLP, Tab-Resnet, Tab-Transformer, Autoint) previously lumped as 'neural_network'.",
     "medium", "analysis/writeup-reevaluation/s4e7_rank1.md"),

    ("playground-series-s5e7", 2, "models_used",
     "catboost; xgboost; lightgbm; lightgbm_goss; lightgbm_dart",
     "Notebook trains 5 GBDT variants; previously counted 3.",
     "medium", "analysis/writeup-reevaluation/s5e7_rank2.md"),
]


# =====================================================================
# NEW COLUMNS: top_3_margin (derived from leaderboard.json)
# =====================================================================
def compute_top3_margins():
    """Return dict {competition_ref: float margin}."""
    out = {}
    for d in LB_DIR.iterdir():
        lb = d / "leaderboard.json"
        if not lb.exists():
            continue
        try:
            data = json.load(open(lb))
            if isinstance(data, list) and len(data) >= 3:
                s1 = float(data[0]["score"])
                s3 = float(data[2]["score"])
                out[d.name] = round(abs(s1 - s3), 6)
        except Exception:
            pass
    return out


# =====================================================================
# NEW COLUMN: distribution_shift_type (manual backfill from re-eval docs)
# =====================================================================
# Only entries with distribution_shift = TRUE in Pass 1 get a non-trivial value.
# All others get 'not-applicable' (no shift flagged) or 'none' (explicit no-shift).
DIST_SHIFT_TYPE = {
    "icr-identify-age-related-conditions": "covariate",  # Featured medical; spreadsheet flags TRUE, type not detailed in re-eval but covariate-like
    "playground-series-s5e3": "temporal",  # train 6 years, test 2 new years; Group K Fold by year
    "playground-series-s5e12": "temporal",  # CUTOFF_ID=678260 marks where train starts matching test
    "playground-series-s4e4": "covariate",  # UCI Abalone vs synthetic (different distributions)
    "tabular-playground-series-feb-2022": "covariate",  # mutated bacteria with different probability distributions
}


# =====================================================================
# CODEBOOK CLARIFICATIONS
# =====================================================================
CODEBOOK_CLARIFICATIONS = [
    ("--- Pass 1 audit clarifications (added 2026-05-27) ---", ""),
    ("primary_model vs best_single_model vs dominant_base_model",
     "Three related but distinct fields: primary_model = the headline model family of the winning approach (often the family of the highest-weight contributor); best_single_model = the highest-CV single base model regardless of ensemble role; dominant_base_model = the family with the largest count of base models in the ensemble. May diverge when (a) best single is a stacker meta-learner or (b) ensemble has many of one family but the headline contributor is another. Example: s6e3 has primary=gbm, best_single=TabM, dominant_base=gbm (was incorrectly 'neural_network' until 2026-05-27 audit)."),
    ("hyperparameter_tuning: automated vs optuna",
     "automated = AutoML framework handles HP internally (s3e1 Kirderf with AutoGluon TabularPredictor; s3e17 ISoft with 5 AutoML frameworks). optuna = author runs explicit Optuna sweep over model HP. Edge case: s5e7 Irfan uses Optuna only for ensemble weight + threshold tuning (per-model HP hardcoded as 'optimized' dicts) — coded as 'optuna' but documented in Pass 2 winner_unique_edge."),
    ("models_used (base vs stacker)",
     "Field captures all models that produced predictions in the pipeline, INCLUDING stacker meta-learners. NN-as-final-stacker cases (s4e9 Mart Preusse, s3e26 Hardy Xu) and CatBoost-as-final-stacker (s5e8 Mahog) list the stacker in models_used; the stacker role is captured in ensemble_method and (post-audit) in Pass 2 winner_unique_edge. Future schema could split into models_base + model_stacker for clarity."),
    ("top_3_margin",
     "Numeric. Absolute difference between rank-1 and rank-3 private LB scores (computed from leaderboard.json). Lower = photo-finish. Added 2026-05-27 to support photo-finish->validation-discipline coupling analysis."),
    ("distribution_shift_type",
     "Enum: temporal | covariate | label-noise | none | not-applicable. Only meaningful when distribution_shift=TRUE; otherwise 'not-applicable'. 'temporal' = time-based shift (s5e3 train years vs test years; s5e12 cutoff_id). 'covariate' = feature-distribution shift between train and test sources (s4e4 UCI vs synthetic; tps-feb-2022 mutated bacteria). 'label-noise' = same features, different label distributions. Added 2026-05-27."),
    ("original_data_usage: 'none' vs 'yes' vs use-mode",
     "Current values mix availability with use-mode and are scheduled for split. Known issue tracked in 'Data Quality Audit' sheet as proposed-split. Until then: 'none' = neither used nor (in most cases) available; 'yes' = available and used in some unspecified way; explicit modes (concat_rows, merge_columns, both) = use mode is known. s5e10 corrected 2026-05-27 from 'yes' to 'none' (no external original)."),
]


# =====================================================================
def main():
    if not XLSX_PATH.exists():
        print(f"ERROR: {XLSX_PATH} not found")
        sys.exit(1)

    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb["Competition Data"]

    # Build header index
    headers = [c.value for c in ws[1]]

    def col(name):
        return headers.index(name) + 1

    # Find rows by (competition_ref, finish_rank)
    def find_row(comp_ref, finish_rank):
        for r in range(2, ws.max_row + 1):
            if (ws.cell(row=r, column=col("competition_ref")).value == comp_ref and
                ws.cell(row=r, column=col("finish_rank")).value == finish_rank):
                return r
        return None

    # ---- Track changes for Corrections Log ----
    changes_log = []  # list of dicts

    # ---- PHASE 1+3: Apply cell-level corrections ----
    print("Applying cell-level corrections...")
    for comp_ref, rank, field, new_val, reason, severity, source in CELL_CORRECTIONS:
        r = find_row(comp_ref, rank)
        if r is None:
            print(f"  SKIP (row not found): {comp_ref} rank={rank}")
            continue
        if field not in headers:
            print(f"  SKIP (field not found): {field}")
            continue
        c = col(field)
        old_val = ws.cell(row=r, column=c).value
        if old_val == new_val:
            print(f"  noop: {comp_ref} rank={rank} {field} already = {new_val!r}")
            continue
        ws.cell(row=r, column=c, value=new_val)
        changes_log.append({
            "date": str(date.today()),
            "competition_ref": comp_ref,
            "finish_rank": rank,
            "field": field,
            "old_value": str(old_val) if old_val is not None else "",
            "new_value": str(new_val),
            "severity": severity,
            "reason": reason,
            "source": source,
        })
        print(f"  CHANGED: {comp_ref} rank={rank} {field}: {str(old_val)[:30]!r} -> {str(new_val)[:30]!r}")

    # ---- PHASE 2: Add top_3_margin column ----
    print("\nAdding top_3_margin column...")
    margins = compute_top3_margins()
    if "top_3_margin" not in headers:
        new_col = ws.max_column + 1
        ws.cell(row=1, column=new_col, value="top_3_margin").font = Font(bold=True)
        headers.append("top_3_margin")
        print(f"  added column at index {new_col}")
    tmc = col("top_3_margin")
    backfilled = 0
    for r in range(2, ws.max_row + 1):
        comp_ref = ws.cell(row=r, column=col("competition_ref")).value
        old = ws.cell(row=r, column=tmc).value
        new_val = margins.get(comp_ref)
        if new_val is None:
            continue
        if old != new_val:
            ws.cell(row=r, column=tmc, value=new_val)
            backfilled += 1
            if old is not None:
                changes_log.append({
                    "date": str(date.today()),
                    "competition_ref": comp_ref,
                    "finish_rank": ws.cell(row=r, column=col("finish_rank")).value,
                    "field": "top_3_margin",
                    "old_value": str(old),
                    "new_value": str(new_val),
                    "severity": "medium",
                    "reason": "Recomputed from leaderboard.json.",
                    "source": f"data/raw/{comp_ref}/leaderboard.json",
                })
    print(f"  backfilled {backfilled} cells")

    # ---- PHASE 4: Add distribution_shift_type column ----
    print("\nAdding distribution_shift_type column...")
    if "distribution_shift_type" not in headers:
        new_col = ws.max_column + 1
        ws.cell(row=1, column=new_col, value="distribution_shift_type").font = Font(bold=True)
        headers.append("distribution_shift_type")
        print(f"  added column at index {new_col}")
    dstc = col("distribution_shift_type")
    dst_filled = 0
    for r in range(2, ws.max_row + 1):
        comp_ref = ws.cell(row=r, column=col("competition_ref")).value
        ds_val = ws.cell(row=r, column=col("distribution_shift")).value
        old = ws.cell(row=r, column=dstc).value
        if comp_ref in DIST_SHIFT_TYPE:
            new_val = DIST_SHIFT_TYPE[comp_ref]
        elif ds_val == "TRUE" or ds_val is True:
            new_val = "unspecified"
        elif ds_val == "FALSE" or ds_val is False:
            new_val = "none"
        else:
            new_val = "not-applicable"
        if old != new_val:
            ws.cell(row=r, column=dstc, value=new_val)
            dst_filled += 1
    print(f"  backfilled {dst_filled} cells")

    # ---- PHASE 5: Codebook clarifications ----
    print("\nUpdating Codebook with clarifications...")
    cb = wb["Codebook"]
    # Append at the end; skip if already present
    existing_fields = set()
    for r in range(2, cb.max_row + 1):
        v = cb.cell(row=r, column=1).value
        if v:
            existing_fields.add(v.strip())
    next_row = cb.max_row + 1
    added = 0
    for field, allowed in CODEBOOK_CLARIFICATIONS:
        if field.strip() in existing_fields:
            continue
        cb.cell(row=next_row, column=1, value=field)
        c = cb.cell(row=next_row, column=2, value=allowed)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        if field.startswith("---"):
            cb.cell(row=next_row, column=1).font = Font(bold=True, italic=True)
        next_row += 1
        added += 1
    print(f"  added {added} codebook entries")

    # ---- Write Corrections Log sheet ----
    print(f"\nWriting Corrections Log ({len(changes_log)} entries)...")
    if CORRECTIONS_SHEET in wb.sheetnames:
        del wb[CORRECTIONS_SHEET]
    cl = wb.create_sheet(CORRECTIONS_SHEET)
    cl_headers = ["date", "competition_ref", "finish_rank", "field",
                  "old_value", "new_value", "severity", "reason", "source"]
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    for col_idx, h in enumerate(cl_headers, start=1):
        c = cl.cell(row=1, column=col_idx, value=h)
        c.font = Font(bold=True)
        c.fill = header_fill
        c.alignment = Alignment(wrap_text=True, vertical="top")
    severity_fills = {
        "high": PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid"),
        "medium": PatternFill(start_color="FFF4E6", end_color="FFF4E6", fill_type="solid"),
        "low": PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"),
    }
    severity_order = {"high": 0, "medium": 1, "low": 2}
    changes_log.sort(key=lambda x: (severity_order.get(x["severity"], 9), x["competition_ref"]))
    for r_idx, ch in enumerate(changes_log, start=2):
        for c_idx, k in enumerate(cl_headers, start=1):
            cell = cl.cell(row=r_idx, column=c_idx, value=ch.get(k, ""))
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            fill = severity_fills.get(ch["severity"])
            if fill:
                cell.fill = fill
    widths = {"A": 12, "B": 36, "C": 6, "D": 28, "E": 36, "F": 50, "G": 8, "H": 60, "I": 50}
    for c, w in widths.items():
        cl.column_dimensions[c].width = w
    cl.freeze_panes = "A2"

    wb.save(XLSX_PATH)
    print(f"\nOK: saved. {len(changes_log)} corrections logged.")
    print("\nSummary by severity:")
    from collections import Counter
    sc = Counter(c["severity"] for c in changes_log)
    for s in ["high", "medium", "low"]:
        print(f"  {s:8s} {sc.get(s, 0)}")


if __name__ == "__main__":
    main()

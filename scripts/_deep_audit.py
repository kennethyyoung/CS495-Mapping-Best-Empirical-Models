"""Deep evaluation of the final dataset — coverage, completeness, field quality."""
import openpyxl
from pathlib import Path
from collections import Counter, defaultdict

XL_PATH = Path(__file__).parent.parent / "data" / "kaggle_meta_analysis.xlsx"
wb = openpyxl.load_workbook(XL_PATH)
ws = wb["Competition Data"]
headers = [c.value for c in ws[1]]

rows = []
for r in range(2, ws.max_row + 1):
    row = {h: ws.cell(r, i+1).value for i, h in enumerate(headers)}
    if row.get("competition_ref"):
        rows.append(row)

BLANK = (None, "", "not_described", "not described")

def filled(v):
    return v not in BLANK

def pct(n, d):
    return f"{n}/{d} ({100*n//d}%)"

print(f"=== DATASET OVERVIEW ===")
print(f"Total entries: {len(rows)}")
print()

# --- Coverage by time period ---
from datetime import datetime
print("=== TEMPORAL COVERAGE ===")
year_counts = Counter()
for r in rows:
    d = r.get("end_date")
    if d:
        try:
            y = d.year if hasattr(d, 'year') else int(str(d)[:4])
            year_counts[y] += 1
        except: pass
for y in sorted(year_counts):
    print(f"  {y}: {year_counts[y]} competitions")
print()

# --- Coverage by target type ---
print("=== TARGET TYPE ===")
for v, c in Counter(r.get("target_type") for r in rows).most_common():
    print(f"  {v}: {c}")
print()

# --- Flowchart field completeness ---
print("=== FLOWCHART FIELD COMPLETENESS ===")
FLOWCHART_FIELDS = [
    ("dominant_base_model",   "Which model type wins?"),
    ("cv_strategy",           "How was CV set up?"),
    ("ensemble_method",       "How were models combined?"),
    ("encoding_strategy",     "How were categoricals encoded?"),
    ("fe_techniques",         "What FE was applied?"),
    ("hyperparameter_tuning", "How was HPO done?"),
    ("original_data_usage",   "Was original data used?"),
    ("scaling",               "Was scaling applied?"),
    ("missing_data_strategy", "How was missing data handled?"),
    ("best_single_model",     "What was the best single model?"),
]
for field, question in FLOWCHART_FIELDS:
    n = sum(1 for r in rows if filled(r.get(field)))
    bar = "#" * n + "." * (len(rows) - n)
    print(f"  {field:<28} {pct(n,len(rows)):>12}  [{bar}]  {question}")
print()

# --- Cross-tab: dominant_base_model vs target_type ---
print("=== dominant_base_model x target_type ===")
ct = defaultdict(Counter)
for r in rows:
    dbm = r.get("dominant_base_model") or "?"
    tt  = r.get("target_type") or "?"
    ct[dbm][tt] += 1
ttypes = ["binary", "regression", "multiclass"]
print(f"  {'':25}", end="")
for tt in ttypes:
    print(f"  {tt:>12}", end="")
print()
for dbm in ["gbm", "neural_network", "linear", "other"]:
    print(f"  {dbm:25}", end="")
    for tt in ttypes:
        print(f"  {ct[dbm][tt]:>12}", end="")
    print()
print()

# --- Cross-tab: dominant_base_model vs has_categorical ---
print("=== dominant_base_model x has_categorical ===")
for dbm in ["gbm", "neural_network", "linear", "other"]:
    cats  = sum(1 for r in rows if r.get("dominant_base_model")==dbm and str(r.get("has_categorical"))=="TRUE")
    nocats= sum(1 for r in rows if r.get("dominant_base_model")==dbm and str(r.get("has_categorical"))=="FALSE")
    print(f"  {dbm:20}  has_cat={cats}  no_cat={nocats}")
print()

# --- n_features distribution ---
print("=== n_features DISTRIBUTION ===")
buckets = Counter()
for r in rows:
    v = r.get("n_features")
    if v and str(v).isdigit():
        n = int(v)
        if n <= 10:   buckets["1-10"] += 1
        elif n <= 20: buckets["11-20"] += 1
        elif n <= 50: buckets["21-50"] += 1
        elif n <= 100:buckets["51-100"] += 1
        else:         buckets["100+"] += 1
for b in ["1-10","11-20","21-50","51-100","100+"]:
    print(f"  {b}: {buckets[b]}")
print()

# --- pct_missing distribution ---
print("=== pct_missing DISTRIBUTION ===")
vals = []
for r in rows:
    v = r.get("pct_missing")
    try:
        vals.append(float(v))
    except: pass
print(f"  0% missing:        {sum(1 for v in vals if v==0)}")
print(f"  0-5% missing:      {sum(1 for v in vals if 0 < v <= 5)}")
print(f"  5-25% missing:     {sum(1 for v in vals if 5 < v <= 25)}")
print(f"  mean pct_missing:  {sum(vals)/len(vals):.2f}%")
print()

# --- Entries where best_single_model is blank ---
print("=== ENTRIES MISSING best_single_model ===")
for r in rows:
    if not filled(r.get("best_single_model")):
        print(f"  {r['competition_ref']:<50}  primary={r.get('primary_model')}  wd={r.get('writeup_detail')}")
print()

# --- Entries where fe_techniques is blank or generic ---
print("=== ENTRIES WITH WEAK fe_techniques ===")
WEAK = (None, "", "not_described", "not described", "None described — AutoML handles internally")
for r in rows:
    v = r.get("fe_techniques")
    if v in WEAK:
        print(f"  {r['competition_ref']:<50}  writeup_detail={r.get('writeup_detail')}")
print()

# --- Fields not in schema worth considering ---
print("=== SCHEMA HEADERS ===")
print("  " + ", ".join(headers))

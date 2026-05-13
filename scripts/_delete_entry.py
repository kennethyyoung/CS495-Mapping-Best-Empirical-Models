import openpyxl
from pathlib import Path

ROOT = Path(__file__).parent.parent
XL_PATH = ROOT / "data" / "kaggle_meta_analysis.xlsx"
SLUG = "tabular-playground-series-nov-2022"

wb = openpyxl.load_workbook(XL_PATH)
ws = wb["Competition Data"]
headers = [c.value for c in ws[1]]
ref_col = headers.index("competition_ref") + 1

for r in range(2, ws.max_row + 1):
    if str(ws.cell(r, ref_col).value) == SLUG:
        ws.delete_rows(r)
        wb.save(XL_PATH)
        print(f"Deleted row for {SLUG}.")
        break
else:
    print(f"{SLUG} not found.")

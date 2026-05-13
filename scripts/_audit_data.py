"""Summarize current state of kaggle_meta_analysis.xlsx."""
import openpyxl
from pathlib import Path
from collections import Counter

XL_PATH = Path(__file__).parent.parent / "data" / "kaggle_meta_analysis.xlsx"

wb = openpyxl.load_workbook(XL_PATH)
ws = wb["Competition Data"]
headers = [c.value for c in ws[1]]

rows = []
for r in range(2, ws.max_row + 1):
    row = {h: ws.cell(r, i+1).value for i, h in enumerate(headers)}
    if row.get("competition_ref"):
        rows.append(row)

print(f"Total entries: {len(rows)}\n")

# Completeness check
BLANK = (None, "", "not_described", "not described")
print("=== BLANK / NOT_DESCRIBED COUNTS ===")
for h in headers:
    blanks = sum(1 for r in rows if r.get(h) in BLANK)
    if blanks > 0:
        print(f"  {h:<35} {blanks:>3} / {len(rows)}")

# Distribution of key categorical fields
KEY_FIELDS = [
    "target_type", "feature_type_dominant", "has_categorical",
    "primary_model", "dominant_base_model", "cv_strategy", "ensemble_method",
    "encoding_strategy", "scaling", "missing_data_strategy",
    "outlier_treatment", "rare_class_handling", "hyperparameter_tuning",
    "original_data_usage", "distribution_shift", "writeup_detail",
]

print("\n=== VALUE DISTRIBUTIONS ===")
for field in KEY_FIELDS:
    counts = Counter(str(r.get(field, "")) for r in rows)
    print(f"\n{field}:")
    for val, cnt in counts.most_common():
        print(f"  {val:<40} {cnt}")

# n_features and pct_missing summary
nf_vals = [r["n_features"] for r in rows if r.get("n_features") not in BLANK]
pm_vals = [r["pct_missing"] for r in rows if r.get("pct_missing") not in BLANK]
print(f"\n=== n_features ({len(nf_vals)} filled) ===")
if nf_vals:
    nf_nums = [int(v) for v in nf_vals]
    print(f"  min={min(nf_nums)}, max={max(nf_nums)}, mean={sum(nf_nums)/len(nf_nums):.1f}")
print(f"\n=== pct_missing ({len(pm_vals)} filled) ===")
pm_nums = []
for r in rows:
    v = r.get("pct_missing")
    if v in BLANK:
        continue
    try:
        pm_nums.append(float(v))
    except (ValueError, TypeError):
        print(f"  BAD VALUE in {r['competition_ref']}: pct_missing={v!r}")
if pm_nums:
    print(f"  min={min(pm_nums)}, max={max(pm_nums):.2f}, mean={sum(pm_nums)/len(pm_nums):.2f}")
    print(f"  entries with any missing: {sum(1 for v in pm_nums if v > 0)}")

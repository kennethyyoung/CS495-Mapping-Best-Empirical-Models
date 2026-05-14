import os
os.chdir(r"C:\Users\Nebi\Documents\capstone\CS495-Mapping-Best-Empirical-Models")
import openpyxl

# Values derived from notebooks (403 blocked the API download):
# s3e4:  V1-V28 + Time + Amount = 30 features, 0% missing (stated in notebook)
# s3e16: Sex + 7 body measurements = 8 features, 0% missing (stated in notebook)
PATCHES = {
    "playground-series-s3e4":  {"n_features": 30,  "pct_missing": 0},
    "playground-series-s3e16": {"n_features": 8,   "pct_missing": 0},
}

wb = openpyxl.load_workbook("data/kaggle_meta_analysis.xlsx")
ws = wb["Competition Data"]
headers = [c.value for c in ws[1]]
ref_col = headers.index("competition_ref") + 1
nf_col  = headers.index("n_features") + 1
pm_col  = headers.index("pct_missing") + 1

for r in range(2, ws.max_row + 1):
    slug = ws.cell(r, ref_col).value
    if slug in PATCHES:
        p = PATCHES[slug]
        if "n_features" in p and ws.cell(r, nf_col).value in (None, "", "not described", "not_described"):
            ws.cell(r, nf_col).value = p["n_features"]
        if "pct_missing" in p and ws.cell(r, pm_col).value in (None, "", "not described", "not_described"):
            ws.cell(r, pm_col).value = p["pct_missing"]
        print(f"Patched {slug}: {p}")

wb.save("data/kaggle_meta_analysis.xlsx")
print("Saved.")

import openpyxl
wb = openpyxl.load_workbook('data/kaggle_meta_analysis.xlsx')
ws = wb['Competition Data']
headers = [c.value for c in ws[1]]
ref_col = headers.index('competition_ref') + 1
refs = [ws.cell(r, ref_col).value for r in range(2, ws.max_row+1)]
s3 = sorted([r for r in refs if r and 's3e' in str(r)])
print(s3)

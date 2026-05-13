"""
Add n_rows, max_cardinality, and metric columns to kaggle_meta_analysis.xlsx.

- n_rows:          number of training rows (from train.csv)
- max_cardinality: highest nunique() across all categorical columns (from train.csv)
- metric:          competition evaluation metric (from Kaggle API metadata)

Skips entries where all three fields are already filled.
"""
import os
import zipfile
import tempfile
from pathlib import Path
import openpyxl
import pandas as pd
from kaggle import KaggleApi

ROOT    = Path(__file__).parent.parent
XL_PATH = ROOT / "data" / "kaggle_meta_analysis.xlsx"

DROP_COLS = {"id", "Id", "ID", "target", "Target", "TARGET"}
CAT_DTYPES = {"object", "category"}

api = KaggleApi()
api.authenticate()

wb = openpyxl.load_workbook(XL_PATH)
ws = wb["Competition Data"]
headers = [c.value for c in ws[1]]

def ensure_col(name):
    if name not in headers:
        idx = len(headers) + 1
        ws.cell(1, idx).value = name
        headers.append(name)
        print(f"  Added column '{name}' at position {idx}")

ensure_col("n_rows")
ensure_col("max_cardinality")
ensure_col("metric")

def col(name):
    return headers.index(name) + 1

BLANK = (None, "", "not_described", "not described")

def already_filled(r):
    return (ws.cell(r, col("n_rows")).value not in BLANK and
            ws.cell(r, col("max_cardinality")).value not in BLANK and
            ws.cell(r, col("metric")).value not in BLANK)

# Cache metric lookups — API call per competition slug
metric_cache = {}

def get_metric(slug):
    if slug in metric_cache:
        return metric_cache[slug]
    try:
        results = api.competitions_list(search=slug)
        match = next((c for c in results if c.ref == slug), None)
        val = str(match.evaluationMetric) if match and match.evaluationMetric else None
        metric_cache[slug] = val
        return val
    except Exception as e:
        print(f"    metric lookup error: {e}")
        metric_cache[slug] = None
        return None

for r in range(2, ws.max_row + 1):
    slug = ws.cell(r, col("competition_ref")).value
    if not slug:
        continue

    if already_filled(r):
        print(f"SKIP  {slug}")
        continue

    print(f"FETCH {slug} ...", end=" ", flush=True)

    # --- metric (no download needed) ---
    if ws.cell(r, col("metric")).value in BLANK:
        m = get_metric(slug)
        if m:
            ws.cell(r, col("metric")).value = m

    # --- n_rows and max_cardinality (need train.csv) ---
    needs_data = (ws.cell(r, col("n_rows")).value in BLANK or
                  ws.cell(r, col("max_cardinality")).value in BLANK)

    if needs_data:
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            downloaded = False
            for filename in ["train.csv", "data.csv"]:
                try:
                    api.competition_download_file(slug, filename, path=tmp, quiet=True)
                    downloaded = True
                    break
                except Exception:
                    pass

            if not downloaded:
                print(f"ERROR: could not download CSV")
                continue

            zips = list(tmp_path.glob("*.zip"))
            if zips:
                with zipfile.ZipFile(zips[0]) as z:
                    z.extractall(tmp_path)

            csvs = list(tmp_path.glob("*.csv"))
            train_csv = next((f for f in csvs if "train" in f.name.lower()), csvs[0] if csvs else None)
            if not train_csv:
                print(f"ERROR: no CSV found")
                continue

            df = pd.read_csv(train_csv)

        # Drop id/target columns
        drop = [c for c in df.columns if c.lower() in {x.lower() for x in DROP_COLS}]
        feat_df = df.drop(columns=drop, errors="ignore")

        n_rows = len(df)
        cat_cols = [c for c in feat_df.columns if str(feat_df[c].dtype) in CAT_DTYPES]
        max_card = int(feat_df[cat_cols].nunique().max()) if cat_cols else 0

        if ws.cell(r, col("n_rows")).value in BLANK:
            ws.cell(r, col("n_rows")).value = n_rows
        if ws.cell(r, col("max_cardinality")).value in BLANK:
            ws.cell(r, col("max_cardinality")).value = max_card

        m_val = ws.cell(r, col("metric")).value
        print(f"n_rows={n_rows}, max_cardinality={max_card}, metric={m_val}")
    else:
        m_val = ws.cell(r, col("metric")).value
        print(f"metric={m_val} (data stats already filled)")

wb.save(XL_PATH)
print("\nDone. Excel saved.")

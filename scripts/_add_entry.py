"""
Adds a new coded entry to kaggle_meta_analysis.xlsx.
Run once per new competition entry.
"""
from pathlib import Path
import openpyxl
from datetime import datetime

ROOT = Path(__file__).parent.parent
XL_PATH = ROOT / "data" / "kaggle_meta_analysis.xlsx"

# --- Entry definition ---
SLUG = "playground-series-s6e4"

ENTRY = {
    "competition_ref":       SLUG,
    "winner_1st":            "kirill0212",
    "winner_2nd":            "cdeotte",
    "winner_3rd":            "nikita7364777",
    "title":                 "Predicting Irrigation Need",
    "category":              "playground",
    "is_monetized":          False,
    "n_teams":               "",
    "end_date":              datetime(2026, 4, 30),
    "target_type":           "multiclass",
    "n_features":            "",
    "pct_missing":           "",
    "feature_type_dominant": "numeric",
    "has_categorical":       False,
    "distribution_shift":    "FALSE",
    "finish_rank":           1,
    "writeup_detail":        2,
    "writeup_url":           "https://www.kaggle.com/competitions/playground-series-s6e4/writeups/1st-place-one-vs-rest-approach",
    "code_url":              "https://www.kaggle.com/code/kirill0212/ps6e4-ensemble-cv-0-98155",
    "missing_data_strategy": "not_described",
    "encoding_strategy":     "not_described",
    "outlier_treatment":     "not_described",
    "scaling":               "not_described",
    "primary_model":         "gbm",
    "cv_strategy":           "stratified_kfold",
    "rare_class_handling":   "none",
    "ensemble_method":       "stacking",
    "fe_techniques":         "not_described",
    "models_used":           "XGBoost, LightGBM, CatBoost, RealMLP, TabM, Logistic Regression",
    "best_single_model":     "XGBoost",
    "hyperparameter_tuning": "not_described",
    "original_data_usage":   "no",
}

# --- Write to Excel ---
wb = openpyxl.load_workbook(XL_PATH)
ws = wb["Competition Data"]

headers = [cell.value for cell in ws[1]]

# Check not already present
existing = [str(ws.cell(row=r, column=headers.index("competition_ref")+1).value)
            for r in range(2, ws.max_row + 1)]
if SLUG in existing:
    print(f"Entry for {SLUG} already exists — skipping.")
else:
    new_row = [ENTRY.get(h, "") for h in headers]
    ws.append(new_row)
    wb.save(XL_PATH)
    print(f"Added entry for {SLUG}.")
    print(f"  Row {ws.max_row}: {len([v for v in new_row if v not in ('', None, 'not described')])} fields filled")

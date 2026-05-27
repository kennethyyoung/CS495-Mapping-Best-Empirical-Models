"""Build 'Data Quality Audit' sheet in kaggle_meta_analysis.xlsx.

Catalogues cell-level issues in the Competition Data sheet surfaced by the
per-writeup re-evaluation. Each row records (competition, field, current
value, issue type, severity, proposed correction, source). Schema-wide
issues (column-level recommendations) are also included with
competition_ref='SCHEMA-WIDE'.

Builds idempotently: deletes existing 'Data Quality Audit' sheet on rerun.
"""

import sys
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from collections import Counter

sys.stdout.reconfigure(encoding='utf-8')

XLSX_PATH = Path("data/kaggle_meta_analysis.xlsx")
SHEET_NAME = "Data Quality Audit"

HEADERS = [
    "competition_ref", "finish_rank", "field", "current_value",
    "issue_type", "severity", "proposed_correction",
    "source_reeval_doc", "notes",
]

# Issue types:
#   inaccurate         — value is factually wrong; should be replaced
#   undercount         — value present but enumeration incomplete
#   miscategorized     — value belongs in a different field
#   semantic-ambiguous — value defensible but field semantics unclear; needs codebook clarification or split
#   missing-data       — value is blank/'not described' but recoverable from another source
#   lumping            — multi-item enumeration collapsed (e.g., 5 NN variants -> 'neural_network')
#   attribution-conflation — field captures 'used' but cannot capture 'originated' (Pass 2 addresses)
#   proposed-add       — schema-wide proposal for new column
#   proposed-split     — schema-wide proposal to split existing column
#   codebook-only      — no data change; codebook needs definitional clarification

# Severity: high / medium / low
#   high   — affects distributional figures in research_report Section 4
#   medium — affects counts but not headline figures
#   low    — interpretive or doesn't affect analysis

AUDIT_ROWS = [
    # ===================== HIGH-SEVERITY CELL-LEVEL =====================
    dict(
        competition_ref="playground-series-s5e2", finish_rank=1, field="n_rows",
        current_value="300000",
        issue_type="inaccurate", severity="high",
        proposed_correction="4000000",
        source_reeval_doc="analysis/writeup-reevaluation/s5e2_rank1.md",
        notes="Combined train is train.csv (300K) + training_extra.csv (~3.7M) loaded together in cdeotte's notebook. Strategy ('single XGB w/ 500 features, no fear of overfitting') is shaped by 4M, not 300K. Misclassifies this entry in any n_rows-based analysis.",
    ),
    dict(
        competition_ref="playground-series-s3e4", finish_rank=3, field="winner_1st",
        current_value="not described",
        issue_type="missing-data", severity="high",
        proposed_correction="Dmitry",
        source_reeval_doc="analysis/writeup-reevaluation/s3e4_rank3.md",
        notes="Actual 1st place per local leaderboard.json. Missing winner names break author-recurrence and community-graph counts.",
    ),
    dict(
        competition_ref="playground-series-s3e4", finish_rank=3, field="winner_2nd",
        current_value="not described",
        issue_type="missing-data", severity="high",
        proposed_correction="Numanynklr",
        source_reeval_doc="analysis/writeup-reevaluation/s3e4_rank3.md",
        notes="Actual 2nd place per local leaderboard.json.",
    ),
    dict(
        competition_ref="playground-series-s6e2", finish_rank=1, field="fe_techniques",
        current_value="...; knowledge distillation (0.5 hard + 0.5 teacher OOF); ...",
        issue_type="miscategorized", severity="high",
        proposed_correction="Remove 'knowledge distillation' from fe_techniques",
        source_reeval_doc="analysis/writeup-reevaluation/s6e2_rank1.md",
        notes="Knowledge distillation is a training-time loss/label technique, not feature engineering. Belongs in models_used or a new training-technique field, not fe_techniques. Note also: writeup §8 lists distillation under 'what did not work' for single-model gain, though it's wired into the published RealMLP notebook for ensemble diversity.",
    ),

    # ===================== MEDIUM-SEVERITY CELL-LEVEL =====================
    dict(
        competition_ref="playground-series-s5e10", finish_rank=1, field="original_data_usage",
        current_value="yes",
        issue_type="semantic-ambiguous", severity="medium",
        proposed_correction="none (or 'internal-augmentation' if codebook extended)",
        source_reeval_doc="analysis/writeup-reevaluation/s5e10_rank1.md",
        notes="Tilii used Lasso to reverse-engineer the generator formula from training data — no external public original dataset was used. 'yes' here conflates internal-data use with external-original use. See SCHEMA-WIDE proposed split of this field.",
    ),
    dict(
        competition_ref="playground-series-s6e3", finish_rank=1, field="dominant_base_model",
        current_value="neural_network",
        issue_type="inaccurate", severity="medium",
        proposed_correction="gbm",
        source_reeval_doc="analysis/writeup-reevaluation/s6e3_rank1.md",
        notes="Writeup count: 90 trees vs 60 NNs — trees dominate by count. NN label likely because best_single_model is TabM. With three columns capturing slightly different things (primary_model / best_single_model / dominant_base_model), codebook should clarify when each may diverge.",
    ),
    dict(
        competition_ref="playground-series-s5e11", finish_rank=1, field="models_used",
        current_value="16 families listed",
        issue_type="undercount", severity="medium",
        proposed_correction="Add: LAMA-DenseLight, LAMA-Dense, TabulaRNN, ExcelFormer, ModernNCA (total 23)",
        source_reeval_doc="analysis/writeup-reevaluation/s5e11_rank1.md",
        notes="Writeup lists 23 distinct architectures; spreadsheet has 16. Affects model-family diversity counts.",
    ),
    dict(
        competition_ref="playground-series-s5e8", finish_rank=2, field="models_used",
        current_value="tabm; xgboost; lightgbm; catboost; neural_network; random_forest; bart (7)",
        issue_type="undercount", severity="medium",
        proposed_correction="Expand to 13: also DeepTables, TabR, RealMLP, Gandalf, GRN, FTTransformer, CNN",
        source_reeval_doc="analysis/writeup-reevaluation/s5e8_rank2.md",
        notes="'neural_network' lumps 7 distinct architectures. Affects NN-architecture-diversity analysis.",
    ),
    dict(
        competition_ref="playground-series-s5e8", finish_rank=2, field="ensemble_method",
        current_value="stacking",
        issue_type="undercount", severity="medium",
        proposed_correction="stacking (CatBoost chosen after head-to-head test of Ridge/LGB/CatB/HC)",
        source_reeval_doc="analysis/writeup-reevaluation/s5e8_rank2.md",
        notes="Mahog explicitly tested 4 stacker families with reported CV/Public LB/Private LB. The chosen CatBoost stacker is a counter-example to the linear-stacker-is-universal pattern. Single 'stacking' value loses the methodologically-interesting head-to-head test.",
    ),
    dict(
        competition_ref="playground-series-s5e10", finish_rank=1, field="models_used",
        current_value="XGBoost; CatBoost; NN; Lasso; AutoGluon; Autoencoder (6)",
        issue_type="undercount", severity="medium",
        proposed_correction="Also: Keras FM, TabM, TF embedding networks (lumped under 'NN')",
        source_reeval_doc="analysis/writeup-reevaluation/s5e10_rank1.md",
        notes="'NN' lumps multiple architectures the writeup names explicitly.",
    ),
    dict(
        competition_ref="playground-series-s5e12", finish_rank=1, field="models_used",
        current_value="XGBoost; LightGBM; CatBoost; NN; Ridge; AutoML",
        issue_type="undercount", severity="medium",
        proposed_correction="Add: HistGradientBoosting (visible in notebook MODELS list)",
        source_reeval_doc="analysis/writeup-reevaluation/s5e12_rank1.md",
        notes="HistGradientBoosting visible in the published notebook's MODELS list. Also note the 'ensemble-of-ensembles' structure (Ridge stacks and prior HC outputs as candidate models inside meta-HC) isn't captured by any field.",
    ),
    dict(
        competition_ref="playground-series-s4e7", finish_rank=1, field="models_used",
        current_value="lightgbm; catboost; xgboost; neural_network (lumped)",
        issue_type="lumping", severity="medium",
        proposed_correction="Expand NN: denselight, mlp, tab_resnet, tab_transformer, autoint",
        source_reeval_doc="analysis/writeup-reevaluation/s4e7_rank1.md",
        notes="Cross Sellers used 5 distinct NN architectures (Denselight, MLP, Tab-Resnet, Tab-Transformer, Autoint) lumped as 'neural_network'.",
    ),
    dict(
        competition_ref="playground-series-s5e7", finish_rank=2, field="models_used",
        current_value="catboost; xgboost; lightgbm (3)",
        issue_type="undercount", severity="medium",
        proposed_correction="Add: LightGBM-GOSS, LightGBM-DART (5 total)",
        source_reeval_doc="analysis/writeup-reevaluation/s5e7_rank2.md",
        notes="Notebook trains 5 GBDT variants; spreadsheet counts 3.",
    ),
    dict(
        competition_ref="playground-series-s5e7", finish_rank=2, field="hyperparameter_tuning",
        current_value="optuna",
        issue_type="semantic-ambiguous", severity="medium",
        proposed_correction="optuna (ensemble-weights + threshold only; per-model HP hardcoded)",
        source_reeval_doc="analysis/writeup-reevaluation/s5e7_rank2.md",
        notes="Optuna used only for ensemble weights + threshold tuning (cell 41); per-model params are hardcoded 'optimized' dicts with no per-model Optuna study in the notebook. Codebook should distinguish 'optuna for HP search' vs 'optuna for ensemble-weight tuning'.",
    ),
    dict(
        competition_ref="playground-series-s4e11", finish_rank=1, field="missing_data_strategy",
        current_value="imputation",
        issue_type="inaccurate", severity="medium",
        proposed_correction="leave_as_is (with per-model native handling)",
        source_reeval_doc="analysis/writeup-reevaluation/s4e11_rank1.md",
        notes="Author explicit in comments: 'usually best to leave the inconsistencies in the data as is.' Notebook does some imputation but the methodological principle is leave-as-is. Codebook value 'leave_as_is' may need to be added if not present.",
    ),
    dict(
        competition_ref="playground-series-s4e10", finish_rank=2, field="best_single_model",
        current_value="not_described",
        issue_type="missing-data", severity="low",
        proposed_correction="(inferable from notebook; author didn't enumerate per-model CV in writeup)",
        source_reeval_doc="analysis/writeup-reevaluation/s4e10_rank2.md",
        notes="Honest 'not_described' is defensible — writeup doesn't enumerate per-model CV scores. Notebook has all 9 base-model CVs computed; could be inferred if needed.",
    ),
    dict(
        competition_ref="playground-series-s6e4", finish_rank=1, field="(own-model count)",
        current_value="6 own models (writeup)",
        issue_type="semantic-ambiguous", severity="low",
        proposed_correction="7 visible in notebook (my_rmlp_xgb_2 is a blend)",
        source_reeval_doc="analysis/writeup-reevaluation/s6e4_rank1.md",
        notes="Writeup says 6 own models; notebook cell 24 shows 7 after exclusions. The discrepancy is one entry being a blend rather than a clean own-model. Not a spreadsheet field per se; flagged for completeness.",
    ),

    # ===================== ATTRIBUTION-CONFLATION (Pass 2 addresses) =====================
    dict(
        competition_ref="SCHEMA-WIDE", finish_rank="", field="fe_techniques",
        current_value="(free text)",
        issue_type="attribution-conflation", severity="medium",
        proposed_correction="No Pass-1 change — Pass 2 'origination_score' + 'cited_community_members' separates 'used' from 'originated'",
        source_reeval_doc="analysis/writeup-reevaluation/INDEX.md",
        notes="Systemic: fe_techniques captures what was used. Re-eval found ~7 entries where 'novel' techniques are credited to named community members (Sergey-s3e14, Bill Cruise-s3e3, Moonlit-s4e3, omid-s4e10, Mart Preusse-s4e9, Kirderf-s3e1, kirill0212-s6e4). Pass 2 sheet resolves cleanly; don't double-encode in Pass 1.",
    ),

    # ===================== SCHEMA-WIDE PROPOSALS =====================
    dict(
        competition_ref="SCHEMA-WIDE", finish_rank="", field="top_3_margin",
        current_value="(does not exist)",
        issue_type="proposed-add", severity="high",
        proposed_correction="Add numeric column: private LB spread between 1st and 3rd",
        source_reeval_doc="analysis/writeup-reevaluation/INDEX.md",
        notes="Photo-finish margin recurs as a strategy-shaping constraint (s4e8 0.00004, s5e10 4-way tie at 0.05563, s6e2 0.00001, s5e12 0.00007). One of 6 promoted couplings (photo-finish -> validation-discipline) cannot be analyzed without this column. Derivable from local leaderboard.json files.",
    ),
    dict(
        competition_ref="SCHEMA-WIDE", finish_rank="", field="distribution_shift_type",
        current_value="(does not exist; only bool distribution_shift)",
        issue_type="proposed-add", severity="medium",
        proposed_correction="Add enum: temporal / covariate / label-noise / none / not-applicable",
        source_reeval_doc="analysis/writeup-reevaluation/INDEX.md",
        notes="Bool flags presence; loses kind. Three known cases each got different handling: s5e3 temporal-by-year (Group K Fold), s5e12 temporal-by-cutoff (post-cutoff CV), s4e4 covariate (train-on-original-no-validate). Cross-case shift-handling analysis requires this column.",
    ),
    dict(
        competition_ref="SCHEMA-WIDE", finish_rank="", field="original_data_usage",
        current_value="enum mixing availability and use mode",
        issue_type="proposed-split", severity="medium",
        proposed_correction="Split into external_original_available (bool) + external_original_use_mode (rows-only / columns-only / both / features-derived / lookup / available-not-used / unavailable / unknown)",
        source_reeval_doc="analysis/writeup-reevaluation/INDEX.md",
        notes="Current 'yes' / 'none' / 'concat_rows' / 'merge_columns' / 'not_described' values smush two distinct questions. cdeotte's s3e5 'available but chose not to use' (CV-driven) and s5e3 'available, used as columns' become distinguishable after the split.",
    ),
    dict(
        competition_ref="SCHEMA-WIDE", finish_rank="", field="primary_model / best_single_model / dominant_base_model",
        current_value="three columns, occasionally divergent",
        issue_type="codebook-only", severity="low",
        proposed_correction="Add codebook entries defining what each captures and when they may diverge",
        source_reeval_doc="analysis/writeup-reevaluation/s6e3_rank1.md",
        notes="s6e3 has primary_model=gbm, best_single_model=TabM, dominant_base_model=neural_network (all different) — the divergence is meaningful but the codebook doesn't define when it should occur. One paragraph of clarification, no data change.",
    ),
    dict(
        competition_ref="SCHEMA-WIDE", finish_rank="", field="hyperparameter_tuning",
        current_value="none / manual / optuna / bayesian / random_search / automated / custom",
        issue_type="codebook-only", severity="low",
        proposed_correction="Codebook: reserve 'automated' for AutoML-handled (s3e17, s3e1); 'optuna' for explicit Optuna sweep over model HP (not ensemble weights only)",
        source_reeval_doc="analysis/writeup-reevaluation/s3e17_rank3.md",
        notes="s5e7 uses Optuna only for ensemble weights + threshold (per-model HP hardcoded) but is currently coded 'optuna' same as competitions with full Optuna model-HP sweeps. The distinction matters when analyzing HP-tuning effort.",
    ),
    dict(
        competition_ref="SCHEMA-WIDE", finish_rank="", field="models_used",
        current_value="(semicolon list, no base/stacker distinction)",
        issue_type="codebook-only", severity="low",
        proposed_correction="Codebook: clarify whether stacker models are included in the list. Note: s4e9 NN-as-stacker, s3e26 NN-as-stacker, s5e8 CatBoost-as-stacker are all currently in the same list as base models.",
        source_reeval_doc="analysis/writeup-reevaluation/s4e9_rank1.md",
        notes="Optional schema extension: separate 'stacker_family' column. Not urgent because ensemble_method partially captures the stacker. Document the limitation in Methodology.",
    ),
]


def main():
    if not XLSX_PATH.exists():
        print(f"ERROR: {XLSX_PATH} not found")
        sys.exit(1)

    wb = openpyxl.load_workbook(XLSX_PATH)
    if SHEET_NAME in wb.sheetnames:
        del wb[SHEET_NAME]
    ws = wb.create_sheet(SHEET_NAME)

    # Header row
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    for col_idx, hdr in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=hdr)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # Data rows — sort schema-wide entries to the top, then cell-level by severity then competition
    severity_order = {"high": 0, "medium": 1, "low": 2}

    def sort_key(r):
        is_schema = 0 if r["competition_ref"] == "SCHEMA-WIDE" else 1
        return (is_schema, severity_order.get(r["severity"], 9), r["competition_ref"])

    sorted_rows = sorted(AUDIT_ROWS, key=sort_key)

    severity_fills = {
        "high": PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid"),
        "medium": PatternFill(start_color="FFF4E6", end_color="FFF4E6", fill_type="solid"),
        "low": PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"),
    }

    for row_idx, row in enumerate(sorted_rows, start=2):
        values = [
            row.get("competition_ref", ""),
            row.get("finish_rank", ""),
            row.get("field", ""),
            row.get("current_value", ""),
            row.get("issue_type", ""),
            row.get("severity", ""),
            row.get("proposed_correction", ""),
            row.get("source_reeval_doc", ""),
            row.get("notes", ""),
        ]
        fill = severity_fills.get(row.get("severity", ""))
        for col_idx, v in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=v)
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            if fill:
                cell.fill = fill

    # Column widths
    widths = {
        "A": 36, "B": 6, "C": 28, "D": 36,
        "E": 22, "F": 8, "G": 60, "H": 44, "I": 75,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"

    wb.save(XLSX_PATH)

    # Print summary
    print(f"OK: wrote sheet '{SHEET_NAME}' with {len(sorted_rows)} audit rows")
    print(f"\nBy severity:")
    sev_counts = Counter(r["severity"] for r in AUDIT_ROWS)
    for s in ["high", "medium", "low"]:
        print(f"  {s:8s} {sev_counts.get(s, 0)}")
    print(f"\nBy issue type:")
    type_counts = Counter(r["issue_type"] for r in AUDIT_ROWS)
    for t, c in sorted(type_counts.items(), key=lambda x: -x[1]):
        print(f"  {t:24s} {c}")
    print(f"\nCell-level vs schema-wide:")
    cell_level = sum(1 for r in AUDIT_ROWS if r["competition_ref"] != "SCHEMA-WIDE")
    schema_wide = sum(1 for r in AUDIT_ROWS if r["competition_ref"] == "SCHEMA-WIDE")
    print(f"  cell-level   {cell_level}")
    print(f"  schema-wide  {schema_wide}")


if __name__ == "__main__":
    main()

"""Fix remaining cosmetic inconsistencies after main normalization pass."""
import openpyxl
from pathlib import Path

XL_PATH = Path(__file__).parent.parent / "data" / "kaggle_meta_analysis.xlsx"

wb = openpyxl.load_workbook(XL_PATH)
ws = wb["Competition Data"]
headers = [c.value for c in ws[1]]

def col(name):
    return headers.index(name) + 1

changes = []

def normalize(r, field, new_val):
    c = col(field)
    old = ws.cell(r, c).value
    if str(old) != str(new_val):
        ws.cell(r, c).value = new_val
        slug = ws.cell(r, col("competition_ref")).value
        changes.append(f"  {slug:<50} {field}: {old!r} -> {new_val!r}")

# Canonical order for multi-value encoding_strategy: target first, then label, then others
ENCODING_CANONICAL_ORDER = [
    "target_encoding", "label_encoding", "ordinal_encoding",
    "one_hot_encoding", "count_encoding",
]

def canonical_encoding(v):
    parts = [p.strip() for p in str(v).replace(",", ";").split(";") if p.strip()]
    ordered = sorted(parts, key=lambda x: ENCODING_CANONICAL_ORDER.index(x)
                     if x in ENCODING_CANONICAL_ORDER else 99)
    return "; ".join(ordered)

for r in range(2, ws.max_row + 1):
    slug = ws.cell(r, col("competition_ref")).value
    if not slug:
        continue

    # encoding_strategy: canonical order for multi-value entries
    v = ws.cell(r, col("encoding_strategy")).value
    if v and ";" in str(v):
        normalize(r, "encoding_strategy", canonical_encoding(v))

    # original_data_usage: comma separator -> semicolon
    v = str(ws.cell(r, col("original_data_usage")).value or "")
    if "," in v:
        normalize(r, "original_data_usage", v.replace(", ", "; ").replace(",", "; "))

    # writeup_detail: "not described" -> "not_described"
    v = ws.cell(r, col("writeup_detail")).value
    if str(v).strip() == "not described":
        normalize(r, "writeup_detail", "not_described")

wb.save(XL_PATH)
print(f"{len(changes)} changes:\n")
for c in changes:
    print(c)
print("\nDone.")

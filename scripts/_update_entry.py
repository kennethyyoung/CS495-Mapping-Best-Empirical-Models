"""
Updates specific fields for an existing entry in kaggle_meta_analysis.xlsx.
Only overwrites fields that are currently not described / empty.
"""
from pathlib import Path
import openpyxl

ROOT = Path(__file__).parent.parent
XL_PATH = ROOT / "data" / "kaggle_meta_analysis.xlsx"

NOT_DESCRIBED = {"not described", "not_described", "none", "nan", "", None}

# --- Define update ---
SLUG = "playground-series-s3e3"

UPDATES = {
    "missing_data_strategy": "none",
    "encoding_strategy": "ordinal",
    "scaling": "standard",
    "outlier_treatment": "none",
    "rare_class_handling": "none",
    "ensemble_method": "weighted_blend",
    "cv_strategy": "stratified_kfold",
    "hyperparameter_tuning": "none",
    "original_data_usage": "concat_rows",
    "code_url": "https://www.kaggle.com/code/bcruise/starting-strong-xgboost-lightgbm-catboost",
}

FORCE_UPDATES = {
    "writeup_url": "https://www.kaggle.com/competitions/playground-series-s3e3/writeups/bill-cruise-1st-place-that-was-unexpected",
}

# --- Apply ---
wb = openpyxl.load_workbook(XL_PATH)
ws = wb["Competition Data"]

headers = [cell.value for cell in ws[1]]
ref_col = headers.index("competition_ref") + 1

target_row = None
for r in range(2, ws.max_row + 1):
    if str(ws.cell(row=r, column=ref_col).value) == SLUG:
        target_row = r
        break

if target_row is None:
    print(f"Entry {SLUG} not found.")
else:
    changed = []
    for field, value in UPDATES.items():
        if field not in headers:
            print(f"  Column '{field}' not in sheet — skipping")
            continue
        col = headers.index(field) + 1
        existing = ws.cell(row=target_row, column=col).value
        if str(existing).strip().lower() in NOT_DESCRIBED:
            ws.cell(row=target_row, column=col).value = value
            changed.append(field)
        else:
            print(f"  Skipping '{field}' — already has value: {existing}")
    for field, value in FORCE_UPDATES.items():
        if field not in headers:
            print(f"  Column '{field}' not in sheet — skipping")
            continue
        col = headers.index(field) + 1
        old = ws.cell(row=target_row, column=col).value
        ws.cell(row=target_row, column=col).value = value
        changed.append(f"{field} (was: {old})")
    wb.save(XL_PATH)
    print(f"Updated {SLUG}: {changed}")

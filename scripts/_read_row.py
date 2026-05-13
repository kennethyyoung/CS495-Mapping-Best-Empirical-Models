import openpyxl
SLUG = "playground-series-s3e17"
wb = openpyxl.load_workbook("data/kaggle_meta_analysis.xlsx")
ws = wb["Competition Data"]
headers = [c.value for c in ws[1]]
ref_col = headers.index("competition_ref") + 1
for r in range(2, ws.max_row + 1):
    if str(ws.cell(r, ref_col).value) == SLUG:
        for h in headers:
            val = ws.cell(r, headers.index(h) + 1).value
            print(f"  {h}: {val}")
        break

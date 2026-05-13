"""
Adds winner_1st, winner_2nd, winner_3rd columns to kaggle_meta_analysis.xlsx
from leaderboard.json files, inserted after competition_ref.
"""
import json
from pathlib import Path
import openpyxl

ROOT = Path(__file__).parent.parent
RAW_DIR = ROOT / "data" / "raw"
XL_PATH = ROOT / "data" / "kaggle_meta_analysis.xlsx"

WINNER_COLS = ["winner_1st", "winner_2nd", "winner_3rd"]


def winner_names(slug: str) -> list[str]:
    path = RAW_DIR / slug / "leaderboard.json"
    if not path.exists():
        return ["", "", ""]
    lb = json.loads(path.read_text(encoding="utf-8"))
    names = [e.get("teamName", "") for e in lb]
    while len(names) < 3:
        names.append("")
    return names[:3]


wb = openpyxl.load_workbook(XL_PATH)
ws = wb["Competition Data"]

headers = {cell.value: cell.column for cell in ws[1]}
ref_col = headers["competition_ref"]

# Insert winner columns after competition_ref if not already present
# Work backwards so column indices stay valid during insertion
for col_name in reversed(WINNER_COLS):
    if col_name not in headers:
        ws.insert_cols(ref_col + 1)
        ws.cell(row=1, column=ref_col + 1).value = col_name
        headers = {cell.value: cell.column for cell in ws[1]}

# Re-read headers after all insertions
headers = {cell.value: cell.column for cell in ws[1]}
ref_col   = headers["competition_ref"]
w1_col    = headers["winner_1st"]
w2_col    = headers["winner_2nd"]
w3_col    = headers["winner_3rd"]

for row in ws.iter_rows(min_row=2):
    slug = str(row[ref_col - 1].value or "").strip()
    if not slug:
        continue
    names = winner_names(slug)
    row[w1_col - 1].value = names[0]
    row[w2_col - 1].value = names[1]
    row[w3_col - 1].value = names[2]

wb.save(XL_PATH)
print("Done.")

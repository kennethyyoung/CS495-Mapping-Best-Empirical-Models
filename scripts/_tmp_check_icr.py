import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl

wb = openpyxl.load_workbook("data/kaggle_meta_analysis.xlsx")
ws = wb["Competition Data"]
headers = [cell.value for cell in ws[1]]
ref_col = headers.index("competition_ref") + 1

fields = ["finish_rank", "code_url", "dominant_base_model", "metric", "n_rows",
          "missing_data_strategy", "encoding_strategy", "scaling",
          "cv_strategy", "ensemble_method", "fe_techniques",
          "hyperparameter_tuning", "original_data_usage"]

for r in range(2, ws.max_row + 1):
    if ws.cell(row=r, column=ref_col).value == "icr-identify-age-related-conditions":
        print(f"ref: icr-identify-age-related-conditions")
        for f in fields:
            if f in headers:
                col = headers.index(f) + 1
                print(f"  {f}: {ws.cell(row=r, column=col).value}")
        break

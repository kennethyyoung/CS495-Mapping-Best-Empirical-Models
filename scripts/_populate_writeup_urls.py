"""
Sets writeup_url to the competition discussion page (sorted by votes) for all entries.
Ensures code_url column exists next to it for manual entry.
"""
from pathlib import Path
import openpyxl

ROOT = Path(__file__).parent.parent
XL_PATH = ROOT / "data" / "kaggle_meta_analysis.xlsx"

wb = openpyxl.load_workbook(XL_PATH)
ws = wb["Competition Data"]

headers = {cell.value: cell.column for cell in ws[1]}
ref_col = headers["competition_ref"]
url_col = headers["writeup_url"]

if "code_url" not in headers:
    ws.insert_cols(url_col + 1)
    ws.cell(row=1, column=url_col + 1).value = "code_url"
    headers = {cell.value: cell.column for cell in ws[1]}

ref_col = headers["competition_ref"]
url_col = headers["writeup_url"]
code_col = headers["code_url"]

for row in ws.iter_rows(min_row=2):
    slug = str(row[ref_col - 1].value or "").strip()
    if not slug:
        continue
    row[url_col - 1].value = (
        f"https://www.kaggle.com/competitions/{slug}/discussion?sortBy=voteCount"
    )
    if not row[code_col - 1].value:
        row[code_col - 1].value = ""

wb.save(XL_PATH)
print("Done.")

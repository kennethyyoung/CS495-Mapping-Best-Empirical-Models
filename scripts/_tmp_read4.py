import os
os.chdir(r"C:\Users\Nebi\Documents\capstone\CS495-Mapping-Best-Empirical-Models")
import openpyxl

wb = openpyxl.load_workbook("data/kaggle_meta_analysis.xlsx")
ws = wb["Competition Data"]
headers = [c.value for c in ws[1]]
ref_col = headers.index("competition_ref") + 1
nf_col  = headers.index("n_features") + 1
pm_col  = headers.index("pct_missing") + 1
tt_col  = headers.index("target_type") + 1

for slug in ["playground-series-s3e4", "playground-series-s3e16", "playground-series-s3e26", "playground-series-s4e12"]:
    for r in range(2, ws.max_row + 1):
        if str(ws.cell(r, ref_col).value) == slug:
            print(f"{slug}: n_features={ws.cell(r, nf_col).value}, pct_missing={ws.cell(r, pm_col).value}, target_type={ws.cell(r, tt_col).value}")
            break

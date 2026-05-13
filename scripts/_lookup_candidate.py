import openpyxl
SLUG = "playground-series-s3e4"
wb = openpyxl.load_workbook("data/kaggle_candidates_v2.xlsx")
ws = wb.active
headers = [c.value for c in ws[1]]
for r in range(2, ws.max_row + 1):
    row = {headers[i]: ws.cell(r, i+1).value for i in range(len(headers))}
    if str(row.get("competition_ref", "")) == SLUG:
        for k, v in row.items():
            print(f"  {k}: {v}")
        break

import os
os.chdir(r"C:\Users\Nebi\Documents\capstone\CS495-Mapping-Best-Empirical-Models")
import openpyxl

# Load all 'keep' slugs from candidates
wb_cand = openpyxl.load_workbook("data/kaggle_candidates_v2.xlsx")
ws_cand = wb_cand.active
headers_c = [c.value for c in ws_cand[1]]
keep = set()
for r in range(2, ws_cand.max_row + 1):
    row = {headers_c[i]: ws_cand.cell(r, i+1).value for i in range(len(headers_c))}
    if str(row.get("recommended_action", "")).strip().lower() == "keep":
        keep.add(str(row["competition_ref"]))

# Load all slugs from meta analysis
wb_meta = openpyxl.load_workbook("data/kaggle_meta_analysis.xlsx")
ws_meta = wb_meta["Competition Data"]
headers_m = [c.value for c in ws_meta[1]]
ref_col = headers_m.index("competition_ref") + 1
in_meta = set(str(ws_meta.cell(r, ref_col).value) for r in range(2, ws_meta.max_row + 1))

missing = keep - in_meta
extra = in_meta - keep

print(f"Keep candidates: {len(keep)}")
print(f"In meta analysis: {len(in_meta)}")
print(f"Missing from meta (in keep but not coded): {sorted(missing) or 'none'}")
print(f"Extra in meta (not in keep): {sorted(extra) or 'none'}")

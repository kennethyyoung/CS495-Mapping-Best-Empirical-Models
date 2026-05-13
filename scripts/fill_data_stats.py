"""
Downloads train.csv for each competition in kaggle_meta_analysis.xlsx,
computes n_features and pct_missing, writes results to the Excel, then
deletes the downloaded file.

Skips entries where both fields are already filled.
"""
import os
import zipfile
import tempfile
from pathlib import Path
import openpyxl
import pandas as pd
from kaggle import KaggleApi

ROOT = Path(__file__).parent.parent
XL_PATH = ROOT / "data" / "kaggle_meta_analysis.xlsx"

TARGET_COLS = ["id", "Id", "ID", "target", "Target", "TARGET"]

api = KaggleApi()
api.authenticate()

wb = openpyxl.load_workbook(XL_PATH)
ws = wb["Competition Data"]
headers = [c.value for c in ws[1]]

ref_col  = headers.index("competition_ref") + 1
nf_col   = headers.index("n_features") + 1
pm_col   = headers.index("pct_missing") + 1
tgt_col  = headers.index("target_type") + 1

def already_filled(r):
    nf = ws.cell(r, nf_col).value
    pm = ws.cell(r, pm_col).value
    return nf not in (None, "", "not_described", "not described") and \
           pm not in (None, "", "not_described", "not described")

for r in range(2, ws.max_row + 1):
    slug = ws.cell(r, ref_col).value
    if not slug:
        continue
    if already_filled(r):
        print(f"SKIP  {slug} (already filled)")
        continue

    print(f"FETCH {slug} ...", end=" ", flush=True)
    with tempfile.TemporaryDirectory() as tmp:
        tmp_path = Path(tmp)
        downloaded = False
        for filename in ["train.csv", "data.csv"]:
            try:
                api.competition_download_file(slug, filename, path=tmp, quiet=True)
                downloaded = True
                break
            except Exception:
                pass
        if not downloaded:
            print(f"ERROR: could not download train.csv or data.csv")
            continue

        # May be zipped
        zips = list(tmp_path.glob("*.zip"))
        if zips:
            with zipfile.ZipFile(zips[0]) as z:
                z.extractall(tmp_path)

        csvs = list(tmp_path.glob("*.csv"))
        train_csv = next((f for f in csvs if "train" in f.name.lower()), csvs[0] if csvs else None)
        if train_csv is None:
            print(f"ERROR: no CSV found in {list(tmp_path.iterdir())}")
            continue

        df = pd.read_csv(train_csv)

    # Drop id/target-like columns to get feature count
    drop_candidates = [c for c in df.columns if c.lower() in [x.lower() for x in TARGET_COLS]]
    # Also drop the last column if it looks like the target (for competition CSVs the label is usually last)
    feature_df = df.drop(columns=drop_candidates, errors="ignore")
    n_features = len(feature_df.columns)

    total_cells = feature_df.size
    missing_cells = feature_df.isnull().sum().sum()
    pct_missing = round(missing_cells / total_cells * 100, 2) if total_cells > 0 else 0

    ws.cell(r, nf_col).value = n_features
    ws.cell(r, pm_col).value = pct_missing
    print(f"n_features={n_features}, pct_missing={pct_missing}%")

wb.save(XL_PATH)
print("\nDone. Excel saved.")

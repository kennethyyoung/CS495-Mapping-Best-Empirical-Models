import openpyxl
from pathlib import Path

ROOT = Path(__file__).parent.parent
XL_PATH = ROOT / "data" / "kaggle_meta_analysis.xlsx"
NOT_DESCRIBED = {"not described", "not_described", "none", "nan", "", None}

wb = openpyxl.load_workbook(XL_PATH)
ws = wb["Competition Data"]
headers = [c.value for c in ws[1]]

target_fields = [
    "fe_techniques", "ensemble_method", "encoding_strategy",
    "missing_data_strategy", "cv_strategy", "original_data_usage",
    "models_used", "best_single_model", "hyperparameter_tuning", "scaling",
]

n_rows = ws.max_row - 1
print(f"Total entries: {n_rows}\n")
print(f"  {'Field':<28} {'Filled':>6}   {'%':>4}")
print("  " + "-" * 44)

for field in target_fields:
    if field not in headers:
        print(f"  {field}: column not found")
        continue
    col = headers.index(field) + 1
    empty = []
    filled = 0
    for r in range(2, ws.max_row + 1):
        val = ws.cell(r, col).value
        slug = ws.cell(r, headers.index("competition_ref") + 1).value
        if str(val).strip().lower() in NOT_DESCRIBED:
            empty.append(slug)
        else:
            filled += 1
    pct = round(filled / n_rows * 100)
    print(f"  {field:<28} {filled:>3}/{n_rows}   {pct:>3}%")
    for s in empty:
        print(f"      - {s}")
